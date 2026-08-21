"""
build_datasets.py
-----------------
Generic, registry-driven dataset ingestion: builds map overlay layers and
writes them to a Supabase-ready import CSV:

    supabase/datasets_import.csv

These rows populate the `public.map_features` table
(dataset, source_id, name, props jsonb, geom geometry(Geometry,4326),
PK (dataset, source_id)); the frontend fetches them by viewport bbox via the
`features_in_bbox` RPC. This mirrors pipeline/build_constraints.py: every
dataset is BEST-EFFORT (a missing/failed source is a warning + skip, never an
abort), every source is env-overridable, and the output geometry is EWKT
('SRID=4326;...') at 6 dp with light simplification for polygons.

Choose which datasets to build with --datasets (comma-separated) or the
DATASETS env var; blank = all. Env override convention: <DATASET-KEY
UPPERCASED>_SRC (e.g. ARTICLE4_SRC), accepting either a URL (downloaded to
data/raw/) or a local file path. A few datasets also honour the shorter
historical names used in their docs (GSP_SRC, TEC_SRC, WATER_SRC,
ONS_RENTS_SRC, LAD_BOUNDARIES_SRC, PTAL_SRC, OSM_POWER_GEOJSON).

REGISTRY (dataset -> default source; all downloads land in data/raw/):

  planning.data.gov.uk group (OGL v3.0, already EPSG:4326 GeoJSON, default
  URL https://files.planning.data.gov.uk/dataset/<slug>.geojson):
    lpa_boundary        local-planning-authority       polygons; props: reference
    local_plan_boundary local-plan-boundary            polygons; props: reference
    article4            article-4-direction-area       polygons; props: reference
    tpo_zone            tree-preservation-zone         polygons; props: reference
    design_code_area    design-code-area               polygons; props: reference

  Student housing:
    uni_campus   learning-provider.data.ac.uk learning-providers-plus.csv.
                 Points; name = provider name; props: ukprn, groups — plus
                 per-provider student stats (students_total / _fulltime /
                 _intl / _pg + intl_pct) joined by UKPRN from an optional
                 HESA_STUDENTS_SRC CSV (see _load_hesa_stats).
    uni_campus_site / uni_building
                 OSM amenity=university grounds + building=university
                 footprints from data/raw/osm_university.geojson (env
                 OSM_UNIVERSITY_GEOJSON), extracted by the workflow from the
                 same Geofabrik UK PBF as the power layers.
    ptal         NO default URL (London Datastore file paths are hashed) —
                 set PTAL_SRC to the TfL PTAL grid CSV (X,Y in EPSG:27700 +
                 PTAL grade / AI). Each grid point becomes a 100 m square.

  Energy:
    power_line / power_substation
                 Built together from data/raw/osm_power.geojson (env
                 OSM_POWER_GEOJSON), the intermediate the WORKFLOW extracts
                 from the Geofabrik UK OSM PBF (env OSM_PBF_SRC) with
                 osmium-tool — see .github/workflows/load-datasets.yml.
                 power=line LineStrings -> power_line (props voltage/operator/
                 cables + parsed props.kv); power=substation points/polygons ->
                 power_substation. power=minor_line is skipped.
    gsp_boundary NESO "GSP regions" zip (Feb 2026 release) — holds the same
                 regions GeoJSON in BNG and WGS84; _maybe_unzip_geo picks the
                 WGS84 member. GSP_SRC overrides for future releases.
    tec_register NO default (set TEC_SRC to a NESO TEC Register CSV export).
                 No geometry of its own: rows are geocoded by fuzzy-joining
                 the connection-site name against power_substation names built
                 in the SAME run; matches emit points at the substation
                 centroid (props: mw, status, customer, site).

  Land / values:
    lad_boundary ONS Open Geography LAD boundaries (ArcGIS FeatureServer,
                 paginated). Emitted whenever the boundaries download works.
    la_rents     ONS_RENTS_SRC (latest PIPR local-authority rents CSV — no
                 stable URL, skip w/ warning if unset) joined by LAD code onto
                 the lad_boundary polygons. props: rent_mean, rent_lq and
                 per-bedroom breakdowns as detected.
    alc          Natural England Provisional ALC (ArcGIS FeatureServer,
                 paginated). Polygons; props: alc_grade.
    water_availability
                 NO default (set WATER_AVAILABILITY_SRC / WATER_SRC to the EA
                 CAMS water resource availability GeoJSON/SHP). Polygons;
                 props: availability colour/status, detected defensively.

Run (or let the Action run it — see .github/workflows/load-datasets.yml):
    python pipeline/build_datasets.py
    python pipeline/build_datasets.py --datasets article4,uni_campus
"""

import argparse
import csv
import datetime as _dt
import hashlib
import json
import os
import re
import shutil
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import pandas as pd
import geopandas as gpd
from shapely import to_wkt
from shapely.geometry import MultiPolygon, Polygon
from shapely.ops import unary_union

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"
OUT = ROOT / "supabase" / "datasets_import.csv"

# Light polygon clean-up (metres, applied in EPSG:27700 then reprojected back).
SIMPLIFY_M = float(os.environ.get("SIMPLIFY_M", "10"))
# Coordinate precision in the output WKT (~0.1 m at 6 dp).
COORD_DP = 6
# ArcGIS FeatureServer page size (servers cap at 2000).
ARCGIS_PAGE = 2000

PLANNING_DATA_BASE = "https://files.planning.data.gov.uk/dataset/"

# dataset-key -> planning.data.gov.uk slug (uniform WGS84 GeoJSON, OGL).
PLANNING_DATASETS = {
    "lpa_boundary":        "local-planning-authority",
    "local_plan_boundary": "local-plan-boundary",
    "article4":            "article-4-direction-area",
    "tpo_zone":            "tree-preservation-zone",
    "design_code_area":    "design-code-area",
}

UNI_CAMPUS_URL = ("https://learning-provider.data.ac.uk/data/"
                  "learning-providers-plus.csv")
# NESO "GSP regions" release (Feb 2026) — a zip holding the regions GeoJSON in
# both British National Grid and WGS84; _maybe_unzip_geo picks the WGS84 one.
GSP_DEFAULT_URL = ("https://api.neso.energy/dataset/"
                   "2810092e-d4b2-472f-b955-d8bea01f9ec0/resource/"
                   "5dfab3dd-f192-40ab-b97f-b365a594293c/download/"
                   "gsp_regions_20260209.zip")
# EA CAMS water availability (Cycle 2). NOTE: this is a signed Azure snapshot
# link that EXPIRED 2026-07-25 — it worked for the initial load; once dead the
# builder just skips with a warning. To refresh: download the zip from
# https://environment.data.gov.uk/dataset/62514eb5-e9d5-4d96-8b73-a40c5b702d43
# and either set WATER_SRC to the fresh signed link (grab it from the browser
# Downloads page) or commit data/raw/water-availability.geojson (the zip,
# renamed — content sniffing handles it).
WATER_DEFAULT_URL = ("https://agrilake2live.file.core.windows.net/gms-datasets/"
                     "d2024f0b-6263-435f-984c-a6ff1a8d5b1b/Water_Resource_"
                     "Availability_and_Abstraction_Reliability_Cycle_2.geojson.zip"
                     "?sv=2022-11-02&se=2026-07-25T11%3A52%3A40Z&sr=f&sp=r"
                     "&sig=H6EqLbVXS2UFyotMZOG98to4kUCurc8WcAJCUQXAZLE%3D")
# UKPN "Grid and Primary sites" (opendatasoft export API — stable URL): every
# UKPN grid/primary substation with its headroom attributes. The DNO's own
# capacity view for London/South East/East.
# TfL PTAL grid via ArcGIS Hub (stable v3 download API): 100 m cell polygons
# with the PTAL_2023 grade + AI, EPSG:4326.
PTAL_DEFAULT_URL = ("https://hub.arcgis.com/api/v3/datasets/"
                    "b932a6039c0f4967a7cb7e3b2b58c1b3_33/downloads/data"
                    "?format=geojson&spatialRefId=4326&where=1%3D1")
UKPN_SITES_URL = ("https://ukpowernetworks.opendatasoft.com/api/explore/v2.1/"
                  "catalog/datasets/grid-and-primary-sites/exports/geojson/"
                  "?lang=en&timezone=Europe%2FLondon")
# NGED (National Grid Electricity Distribution) substation capacity CSV from
# their Connected Data portal (CKAN resource URL — stable).
NGED_SITES_URL = ("https://connecteddata.nationalgrid.co.uk/dataset/"
                  "967404e0-f25c-469b-8857-1a396f3c363f/resource/"
                  "d1895bd3-d9d2-4886-a0a3-b7eadd9ab6c2/download/"
                  "substations.csv?format=csv")
# The four remaining DNOs, completing national headroom coverage alongside
# UKPN + NGED. All run opendatasoft portals (except SSEN) whose dataset slugs
# move around — the defaults below are best guesses; when one 404s the run
# warns and the *_SITES_SRC repo variable takes an updated export URL
# (see docs/MANUAL_TASKS.md 9b).
DNO_SITES = [
    # SPEN publishes DG heat maps per licence area — primaries carry a
    # green/amber/red capacity class. "|" separates multiple export URLs;
    # the builder concatenates them into one layer.
    ("spen_sites", "SPEN_SITES_SRC",
     "https://spenergynetworks.opendatasoft.com/api/explore/v2.1/catalog/"
     "datasets/distributed-generation-sp-distribution-heat-maps-spd-primary-"
     "substations/exports/csv?delimiter=%2C"
     "|https://spenergynetworks.opendatasoft.com/api/explore/v2.1/catalog/"
     "datasets/distributed-generation-sp-manweb-heat-maps-spm-primary-"
     "substations/exports/csv?delimiter=%2C",
     "spen-sites.csv", "SP Energy Networks (Scotland S + Merseyside/N Wales)"),
    ("npg_sites", "NPG_SITES_SRC",
     "https://northernpowergrid.opendatasoft.com/api/explore/v2.1/catalog/"
     "datasets/heatmapsubstationareas/exports/csv?delimiter=%2C",
     "npg-sites.csv", "Northern Powergrid (North East + Yorkshire)"),
    ("enwl_sites", "ENWL_SITES_SRC",
     "https://electricitynorthwest.opendatasoft.com/api/explore/v2.1/catalog/"
     "datasets/distribution-tx-headroom/exports/csv?delimiter=%2C",
     "enwl-sites.csv", "Electricity North West"),
    ("ssen_sites", "SSEN_SITES_SRC",
     None,   # data.ssen.co.uk uses hashed CKAN links — needs the var.
     "ssen-sites.csv", "SSEN (north Scotland + central southern England)"),
]
LAD_DEFAULT_URL = ("https://services1.arcgis.com/ESMARspQHYMw9BZ9/arcgis/rest/"
                   "services/Local_Authority_Districts_December_2024_Boundaries"
                   "_UK_BGC/FeatureServer/0/query?where=1%3D1&outFields=*"
                   "&outSR=4326&f=geojson")
# MSOA 2021 generalised boundaries, same ONS Open Geography org. Slug is
# best-effort (unverifiable from this sandbox); MSOA_BOUNDARIES_SRC overrides.
MSOA_DEFAULT_URL = ("https://services1.arcgis.com/ESMARspQHYMw9BZ9/arcgis/rest/"
                    "services/Middle_layer_Super_Output_Areas_December_2021"
                    "_Boundaries_EW_BGC_V3/FeatureServer/0/query?where=1%3D1"
                    "&outFields=*&outSR=4326&f=geojson")
# Natural England ALC via a signed ArcGIS Hub export link (EXPIRED
# 2026-07-24T13:01Z — used for the initial load; once dead the builder skips
# with a warning). To refresh: naturalengland-defra.opendata.arcgis.com →
# Provisional ALC → Download → GeoJSON → copy the link (from the browser
# Downloads page, Ctrl+J) and set ALC_SRC, or hand the file to Claude.
ALC_DEFAULT_URL = ("https://stg-arcgisazurecdataprod.az.arcgis.com/"
                   "exportfiles-3106-14332/Provisional%20Agricultural%20Land%20"
                   "Classification%20%28ALC%29%20%28England%29_"
                   "-752430394565312091.geojson?sv=2025-05-05"
                   "&st=2026-07-24T11%3A56%3A31Z&se=2026-07-24T13%3A01%3A31Z"
                   "&sr=b&sp=r&sig=NJ05rKGC58pUkAzzRSm9OiBqUuvrFOr9S4peYGBGFus%3D")

ALL_DATASETS = [
    "lpa_boundary", "local_plan_boundary", "article4", "tpo_zone",
    "design_code_area",
    "uni_campus", "uni_campus_site", "uni_building", "ptal",
    "power_line", "power_substation", "gsp_boundary", "tec_register",
    "ukpn_sites", "nged_sites",
    "spen_sites", "npg_sites", "enwl_sites", "ssen_sites",
    "lad_boundary", "la_rents", "alc", "water_availability", "hdt",
    "planit_rates", "bus_route",
    "ofcom_fibre", "census_students", "student_accom",
    "building_height",
    "build_cost_index", "lad_income", "msoa_income", "land_value",
    "cil_rates", "lsoa_boundary", "local_plan_housing", "housing_need",
    "connectivity_lsoa", "connectivity_lad", "connectivity_oa",
]

# dataset -> {"count": int|None, "reason": str} filled in as the run proceeds.
STATUS = {}
# (name, centroid geometry EPSG:4326) for every substation built this run —
# tec_register geocodes against this.
_SUBSTATIONS = []


def _note(dataset, reason, count=None):
    STATUS[dataset] = {"count": count, "reason": reason}


def _warn(dataset, msg):
    print(f"  [{dataset}] WARNING: {msg}")


# --- generic helpers ---------------------------------------------------------

ID_CANDIDATES = [
    "source_id", "reference", "entity", "id", "gml_id", "uid", "fid",
    "objectid", "ogc_fid", "@id",
]
NAME_CANDIDATES = ["name", "site_name", "sitename", "title", "label"]


def _find_col(df, candidates, contains=False):
    """Case-insensitive lookup of the first candidate column present. With
    contains=True, also accept a column whose name merely CONTAINS a
    candidate (defensive matching for messy CSV headers)."""
    cols = [c for c in df.columns if c != "geometry"]
    lower = {str(c).strip().lower(): c for c in cols}
    for cand in candidates:
        if cand.lower() in lower:
            return lower[cand.lower()]
    if contains:
        for cand in candidates:
            for lc, orig in lower.items():
                if cand.lower() in lc:
                    return orig
    return None


def _jsonable(v):
    """Coerce a cell value (possibly a numpy scalar) to JSON-serialisable."""
    if v is None:
        return None
    if isinstance(v, (str, bool, int, float)):
        return v
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    item = getattr(v, "item", None)
    if callable(item):
        try:
            return item()
        except Exception:
            pass
    return str(v)


def _cell(feat, col):
    """A cleaned string value for a feature column, or '' when absent/NaN."""
    if not col or col not in feat:
        return ""
    v = feat[col]
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    return str(v).strip()


def _num(feat, col):
    """A float value for a column, or None."""
    s = _cell(feat, col)
    if not s:
        return None
    try:
        return float(str(s).replace(",", ""))
    except ValueError:
        return None


def _stable_id(dataset, ewkt):
    """Deterministic id from dataset+geometry, for sources with no feature id
    — re-runs upsert the same rows instead of duplicating."""
    h = hashlib.md5(f"{dataset}|{ewkt}".encode("utf-8")).hexdigest()[:16]
    return f"{dataset}-{h}"


def _slug_key(s):
    """A safe snake_case props key from a messy CSV header."""
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", str(s).lower())).strip("_")


def _read_csv(path):
    """CSV read tolerant of BOMs and legacy encodings."""
    for enc in ("utf-8-sig", "latin-1"):
        try:
            return pd.read_csv(path, encoding=enc, low_memory=False)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, low_memory=False)


# --- source resolution (URL or local path, env-overridable) ------------------

def _download(url, dest):
    """Download url -> dest (streamed). Raises on failure; callers treat any
    exception as skip-with-warning."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "mastermapper-pipeline/1.0"})
    tmp = dest.with_suffix(dest.suffix + ".part")
    with urllib.request.urlopen(req, timeout=300) as resp, tmp.open("wb") as fh:
        while True:
            chunk = resp.read(1 << 20)
            if not chunk:
                break
            fh.write(chunk)
    tmp.replace(dest)
    return dest


def _resolve_source(dataset, env_vars, default_url, cache_name):
    """Resolve a dataset source into a local file path (or None + reason).

    env_vars: env var names checked in order (the generic <KEY>_SRC first).
    The value may be a URL (downloaded to data/raw/cache_name) or a local
    path. When no env var is set, an existing data/raw/cache_name is reused,
    else default_url is downloaded (best-effort). Returns (Path|None, reason).
    """
    for var in env_vars:
        v = os.environ.get(var, "").strip()
        if not v:
            continue
        if v.lower().startswith(("http://", "https://")):
            dest = RAW / cache_name
            try:
                print(f"  [{dataset}] downloading {var} -> {dest.name} ...")
                return _download(v, dest), f"downloaded from {var}"
            except Exception as exc:
                return None, f"{var} download failed: {exc}"
        p = Path(v)
        if p.exists():
            return p, f"local file via {var}"
        return None, f"{var}={v} does not exist"

    cached = RAW / cache_name
    if cached.exists() and cached.stat().st_size > 0:
        return cached, f"cached {cached.name}"

    if default_url:
        try:
            print(f"  [{dataset}] downloading default source -> {cached.name} ...")
            return _download(default_url, cached), "downloaded default"
        except urllib.error.HTTPError as exc:
            return None, f"default URL returned HTTP {exc.code}"
        except Exception as exc:
            return None, f"default download failed: {exc}"
    return None, "no source configured"


def _fetch_arcgis(url, dest, dataset):
    """Fetch every feature from an ArcGIS FeatureServer query URL (which caps
    responses at ~2000 features) by looping resultOffset, and write one merged
    GeoJSON FeatureCollection to dest. Raises on failure."""
    feats, offset = [], 0
    sep = "&" if "?" in url else "?"
    for _page in range(500):  # hard guard: 500 pages = 1M features
        page_url = (f"{url}{sep}resultOffset={offset}"
                    f"&resultRecordCount={ARCGIS_PAGE}")
        req = urllib.request.Request(
            page_url, headers={"User-Agent": "mastermapper-pipeline/1.0"})
        with urllib.request.urlopen(req, timeout=300) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if "error" in data:
            raise RuntimeError(f"ArcGIS error: {data['error']}")
        page = data.get("features", [])
        feats.extend(page)
        print(f"  [{dataset}] ArcGIS page at offset {offset}: {len(page)} features")
        more = (data.get("exceededTransferLimit")
                or data.get("properties", {}).get("exceededTransferLimit"))
        if not page or (len(page) < ARCGIS_PAGE and not more):
            break
        offset += len(page)
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("w", encoding="utf-8") as fh:
        json.dump({"type": "FeatureCollection", "features": feats}, fh)
    return dest


def _resolve_arcgis_source(dataset, env_vars, default_url, cache_name):
    """Like _resolve_source but understands paginated ArcGIS query URLs (env
    value or default). Plain URLs/paths fall through to _resolve_source."""
    for var in env_vars:
        v = os.environ.get(var, "").strip()
        if v and ("featureserver" in v.lower() or "mapserver" in v.lower()):
            dest = RAW / cache_name
            try:
                print(f"  [{dataset}] fetching paginated ArcGIS source ({var}) ...")
                return _fetch_arcgis(v, dest, dataset), f"ArcGIS via {var}"
            except Exception as exc:
                return None, f"{var} ArcGIS fetch failed: {exc}"
    env_set = any(os.environ.get(v, "").strip() for v in env_vars)
    if env_set:
        return _resolve_source(dataset, env_vars, None, cache_name)
    cached = RAW / cache_name
    if cached.exists() and cached.stat().st_size > 0:
        return cached, f"cached {cached.name}"
    if default_url:
        try:
            print(f"  [{dataset}] fetching paginated ArcGIS default ...")
            return _fetch_arcgis(default_url, cached, dataset), "ArcGIS default"
        except urllib.error.HTTPError as exc:
            return None, f"default ArcGIS URL returned HTTP {exc.code}"
        except Exception as exc:
            return None, f"default ArcGIS fetch failed: {exc}"
    return None, "no source configured"


# --- geometry + row emission -------------------------------------------------

def _to_multipolygon(geom):
    """A non-empty MultiPolygon from any polygonal input, or None."""
    if geom is None or geom.is_empty:
        return None
    gt = geom.geom_type
    if gt == "Polygon":
        return MultiPolygon([geom])
    if gt == "MultiPolygon":
        return geom
    if gt == "GeometryCollection":
        polys = [g for g in geom.geoms if g.geom_type in ("Polygon", "MultiPolygon")]
        if not polys:
            return None
        return _to_multipolygon(unary_union(polys))
    return None


def _ensure_4326(gdf, assume_epsg=4326):
    if gdf.crs is None:
        gdf = gdf.set_crs(assume_epsg)
    try:
        if gdf.crs.to_epsg() != 4326:
            gdf = gdf.to_crs(4326)
    except Exception:
        gdf = gdf.to_crs(4326)
    return gdf


def _emit(gdf, dataset, name_col=None, id_col=None, props_fn=None,
          want="polygon", assume_epsg=4326, simplify=True):
    """Shared tail: normalise CRS to 4326, lightly simplify polygons (in
    EPSG:27700, SIMPLIFY_M metres), and emit one row dict per feature with
    EWKT geometry at 6 dp. want: 'polygon' (promote to MultiPolygon, drop
    non-polygonal), 'point', 'line', or 'any'."""
    if gdf is None or gdf.empty:
        return []
    gdf = gdf[gdf.geometry.notna() & ~gdf.geometry.is_empty].copy()
    if gdf.empty:
        return []
    try:
        gdf["geometry"] = gdf.geometry.make_valid()
    except Exception:
        pass
    gdf = _ensure_4326(gdf, assume_epsg=assume_epsg)

    if want == "polygon":
        gdf = gdf[gdf.geometry.geom_type.isin(["Polygon", "MultiPolygon",
                                               "GeometryCollection"])]
    elif want == "point":
        gdf = gdf[gdf.geometry.geom_type.isin(["Point", "MultiPoint"])]
    elif want == "line":
        gdf = gdf[gdf.geometry.geom_type.isin(["LineString", "MultiLineString"])]
    if gdf.empty:
        return []

    # Light simplification for polygonal (and line) geometry: round-trip
    # through British National Grid so the tolerance is in metres.
    if simplify and SIMPLIFY_M and want in ("polygon", "line", "any"):
        try:
            g27 = gdf.to_crs(27700)
            g27["geometry"] = g27.geometry.simplify(SIMPLIFY_M,
                                                    preserve_topology=True)
            gdf = g27.to_crs(4326)
            gdf = gdf[gdf.geometry.notna() & ~gdf.geometry.is_empty]
        except Exception as exc:
            print(f"  [{dataset}] simplify skipped ({exc})")

    if id_col is None:
        id_col = _find_col(gdf, ID_CANDIDATES)
    if name_col is None:
        name_col = _find_col(gdf, NAME_CANDIDATES)

    rows, seen = [], {}
    for _, feat in gdf.iterrows():
        geom = feat.geometry
        if geom is None or geom.is_empty:
            continue
        if want == "polygon":
            geom = _to_multipolygon(geom)
            if geom is None or geom.is_empty:
                continue
        ewkt = "SRID=4326;" + to_wkt(geom, rounding_precision=COORD_DP, trim=True)

        sid = _cell(feat, id_col)
        if not sid:
            sid = _stable_id(dataset, ewkt)
        # (dataset, source_id) is the PK — make repeats unique deterministically
        # so one upsert batch never touches the same row twice.
        seen[sid] = seen.get(sid, 0) + 1
        if seen[sid] > 1:
            sid = f"{sid}-{seen[sid]}"

        props = {}
        if props_fn is not None:
            for k, v in (props_fn(feat) or {}).items():
                v = _jsonable(v)
                if v not in (None, ""):
                    props[k] = v

        rows.append({
            "dataset": dataset,
            "source_id": sid,
            "name": _cell(feat, name_col) or None,
            "props": props,
            "geom_wkt": ewkt,
        })
    return rows


# --- builders ----------------------------------------------------------------
# Each builder returns {dataset: rows-list} for the datasets it produced, and
# records a STATUS reason for anything it skipped.

def build_planning_dataset(key):
    slug = PLANNING_DATASETS[key]
    path, how = _resolve_source(key, [f"{key.upper()}_SRC"],
                                PLANNING_DATA_BASE + f"{slug}.geojson",
                                f"{slug}.geojson")
    if path is None:
        _warn(key, f"{how} — set {key.upper()}_SRC to a "
                   f"planning.data.gov.uk {slug} GeoJSON (URL or path)")
        _note(key, how)
        return {}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    gdf = gpd.read_file(path)
    ref_col = _find_col(gdf, ["reference"])
    rows = _emit(gdf, key, want="polygon",
                 props_fn=lambda f: {"reference": _cell(f, ref_col)})
    _note(key, how, len(rows))
    return {key: rows}


def _load_hesa_stats():
    """Optional per-provider student stats keyed by UKPRN, merged onto the
    uni_campus points for the PBSA deep-dive card. Source: HESA_STUDENTS_SRC —
    a CSV with a UKPRN column plus any of (detected case-insensitively,
    'contains' matching): total students, full-time, international/non-UK,
    postgraduate. Wide format, one row per provider; trim HESA's metadata
    preamble lines if the raw download has them. Missing source = quiet skip
    (the layer still builds, the card just has no stats)."""
    path, how = _resolve_source("uni_campus_stats", ["HESA_STUDENTS_SRC"],
                                None, "hesa_students.csv")
    if path is None:
        print("  [uni_campus] no HESA_STUDENTS_SRC — points build without "
              "student stats (set it to a per-provider CSV keyed by UKPRN)")
        return {}
    try:
        df = _read_csv(path)
    except Exception as exc:
        _warn("uni_campus", f"HESA stats CSV unreadable ({exc}) — continuing "
                            "without stats")
        return {}
    ukprn_col = _find_col(df, ["ukprn"], contains=True)
    if ukprn_col is None:
        _warn("uni_campus", "HESA stats CSV has no UKPRN column — continuing "
                            f"without stats (columns: {list(df.columns)[:10]})")
        return {}
    fields = {
        "students_total": _find_col(df, ["total_students", "total students",
                                         "students_total", "all students",
                                         "total"], contains=True),
        "students_fulltime": _find_col(df, ["full-time", "fulltime", "full time"],
                                       contains=True),
        "students_intl": _find_col(df, ["international", "non-uk", "non uk",
                                        "overseas"], contains=True),
        "students_pg": _find_col(df, ["postgraduate", "pg "], contains=True),
        # Term-time accommodation mix (HESA DT051 Table 57, full-time
        # students) — the core PBSA demand signals: private halls = existing
        # PBSA penetration, other rented = the HMO market, provider halls =
        # university-run stock.
        "acc_provider_halls": _find_col(df, ["acc_provider_halls",
                                             "provider maintained"], contains=True),
        "acc_private_halls": _find_col(df, ["acc_private_halls",
                                            "private-sector halls",
                                            "private sector halls"], contains=True),
        "acc_parental_home": _find_col(df, ["acc_parental_home", "parental"],
                                       contains=True),
        "acc_own_residence": _find_col(df, ["acc_own_residence", "own residence"],
                                       contains=True),
        "acc_other_rented": _find_col(df, ["acc_other_rented", "other rented"],
                                      contains=True),
    }
    found = {k: c for k, c in fields.items() if c is not None}
    if not found:
        _warn("uni_campus", "no recognisable student-count columns in the HESA "
                            "CSV — continuing without stats")
        return {}
    stats = {}
    for _, r in df.iterrows():
        try:
            ukprn = str(int(float(r[ukprn_col])))
        except (TypeError, ValueError):
            continue
        row = {}
        for k, col in found.items():
            v = pd.to_numeric(pd.Series([r[col]]).astype(str)
                              .str.replace(",", "", regex=False),
                              errors="coerce").iloc[0]
            if pd.notna(v):
                row[k] = int(v)
        if row:
            # Derived shares the PBSA card leads with.
            if "students_total" in row and row["students_total"] > 0 \
                    and "students_intl" in row:
                row["intl_pct"] = round(100.0 * row["students_intl"]
                                        / row["students_total"], 1)
            ft = row.get("students_fulltime")
            if ft:
                for acc, pct in (("acc_private_halls", "pbsa_pct"),
                                 ("acc_provider_halls", "uni_halls_pct"),
                                 ("acc_other_rented", "rented_pct"),
                                 ("acc_parental_home", "home_pct")):
                    if acc in row:
                        row[pct] = round(100.0 * row[acc] / ft, 1)
            stats[ukprn] = row
    print(f"  [uni_campus] HESA stats joined for {len(stats)} providers ({how})")
    return stats


def build_uni_campus():
    key = "uni_campus"
    path, how = _resolve_source(key, ["UNI_CAMPUS_SRC"], UNI_CAMPUS_URL,
                                "learning-providers-plus.csv")
    if path is None:
        _warn(key, f"{how} — set UNI_CAMPUS_SRC to the learning-providers-plus"
                   " CSV (URL or path)")
        _note(key, how)
        return {}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    df = _read_csv(path)
    lat_col = _find_col(df, ["latitude", "lat"], contains=True)
    lon_col = _find_col(df, ["longitude", "long", "lon"], contains=True)
    name_col = _find_col(df, ["provider_name", "view_name", "name"], contains=True)
    ukprn_col = _find_col(df, ["ukprn"], contains=True)
    groups_col = _find_col(df, ["groups"], contains=True)
    if lat_col is None or lon_col is None:
        _warn(key, f"no LATITUDE/LONGITUDE columns found in {path.name} "
                   f"(columns: {list(df.columns)[:12]}...)")
        _note(key, "no coordinate columns in CSV")
        return {}
    df = df.copy()
    df["_lat"] = pd.to_numeric(df[lat_col], errors="coerce")
    df["_lon"] = pd.to_numeric(df[lon_col], errors="coerce")
    before = len(df)
    df = df[df["_lat"].notna() & df["_lon"].notna()]
    if before - len(df):
        print(f"  [{key}] dropped {before - len(df)} row(s) without coordinates")
    gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df["_lon"], df["_lat"]),
                           crs=4326)
    hesa = _load_hesa_stats()

    def _props(f):
        ukprn_raw = _cell(f, ukprn_col)
        p = {"ukprn": ukprn_raw, "groups": _cell(f, groups_col)}
        try:
            p.update(hesa.get(str(int(float(ukprn_raw))), {}))
        except (TypeError, ValueError):
            pass
        return p

    rows = _emit(gdf, key, name_col=name_col, id_col=ukprn_col, want="point",
                 props_fn=_props)
    _note(key, how + (" + HESA stats" if hesa else ""), len(rows))
    return {key: rows}


def build_university_group():
    """uni_campus_site (amenity=university grounds) + uni_building
    (building=university footprints) from the workflow's osmium university
    extract — the full physical footprint of each institution, complementing
    the one-dot-per-provider uni_campus layer."""
    src = os.environ.get("OSM_UNIVERSITY_GEOJSON", "").strip() \
        or str(RAW / "osm_university.geojson")
    path = Path(src)
    if not path.exists():
        for k in ("uni_campus_site", "uni_building"):
            _note(k, "osm_university.geojson missing")
        _warn("uni_campus_site/uni_building",
              f"{path} not found. The GitHub workflow extracts it from the UK"
              " OSM PBF alongside the power extract; locally, run the osmium"
              " steps from .github/workflows/load-datasets.yml or set"
              " OSM_UNIVERSITY_GEOJSON.")
        return {}
    print(f"  [university] reading {path.name} ...")
    gdf = gpd.read_file(path)
    amenity_col = _find_col(gdf, ["amenity"])
    building_col = _find_col(gdf, ["building"])
    name_col = _find_col(gdf, ["name"])
    op_col = _find_col(gdf, ["operator"])
    web_col = _find_col(gdf, ["website"])
    out = {}

    is_poly = gdf.geometry.geom_type.isin(["Polygon", "MultiPolygon"])
    if amenity_col is not None:
        am = gdf[amenity_col].astype(str).str.strip().str.lower()
        sites = gdf[(am == "university") & is_poly]
        out["uni_campus_site"] = _emit(
            sites, "uni_campus_site", name_col=name_col, want="polygon",
            props_fn=lambda f: {"operator": _cell(f, op_col),
                                "website": _cell(f, web_col)})
        _note("uni_campus_site", "osmium extract", len(out["uni_campus_site"]))
    else:
        _note("uni_campus_site", "no amenity tag in extract")

    if building_col is not None:
        bl = gdf[building_col].astype(str).str.strip().str.lower()
        builds = gdf[(bl == "university") & is_poly]
        out["uni_building"] = _emit(
            builds, "uni_building", name_col=name_col, want="polygon",
            props_fn=lambda f: {"operator": _cell(f, op_col)})
        _note("uni_building", "osmium extract", len(out["uni_building"]))
    else:
        _note("uni_building", "no building tag in extract")
    return out


def build_ptal():
    key = "ptal"
    path, how = _resolve_source(key, ["PTAL_SRC"], PTAL_DEFAULT_URL,
                                "ptal_grid.csv")
    if path is None:
        _warn(key, f"{how} — set PTAL_SRC to a PTAL grid vector file "
                   "(GeoJSON/shapefile zip) or an X/Y CSV.")
        _note(key, how)
        return {}
    # The Datastore offers the grid BOTH as an attribute CSV (needs X/Y
    # columns) and as vector files (shapefile zip / GeoJSON with the PTAL
    # attribute on real 100 m cell polygons). Accept either: a zip or vector
    # file takes the vector path; only a genuine CSV takes the X/Y path.
    vec = _maybe_unzip_geo(path, key, prefer=("ptal",))
    if vec is None:
        _note(key, "zip source held no vector file")
        return {}
    # Content sniff: the download cache is always named ptal_grid.csv, so a
    # GeoJSON fetched from the ArcGIS Hub default URL arrives with a .csv
    # suffix — route by first byte, not extension.
    try:
        head = vec.open("rb").read(64).lstrip()
    except OSError:
        head = b""
    is_json = head.startswith(b"{")
    if is_json or vec != path or vec.suffix.lower() in (".geojson", ".json", ".gpkg", ".shp"):
        print(f"  [{key}] reading vector source {vec.name} ({how}) ...")
        # gpd.read_file picks its OGR driver from the file EXTENSION, so
        # GeoJSON cached under the .csv name gets parsed by the CSV driver
        # (columns come out as ['{...']). Alias it to .geojson first.
        read_target = vec
        if is_json and vec.suffix.lower() not in (".geojson", ".json", ".gpkg"):
            alias = vec.with_suffix(".sniffed.geojson")
            shutil.copyfile(vec, alias)
            read_target = alias
        gdf = gpd.read_file(read_target)
        ptal_col = _find_col(gdf, ["ptal"], contains=True)
        ai_col = _find_col(gdf, ["ai", "access index"], contains=False) \
            or _find_col(gdf, ["access index"], contains=True)
        if ptal_col is None:
            _warn(key, f"no PTAL column in {vec.name} "
                       f"(columns: {list(gdf.columns)[:10]}...)")
            _note(key, "no PTAL column in vector source")
            return {}
        rows = _emit(gdf, key, want="any",
                     props_fn=lambda f: {"ptal": _cell(f, ptal_col),
                                         "ai": _num(f, ai_col) if ai_col else None})
        _note(key, how, len(rows))
        return {key: rows}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    df = _read_csv(path)
    x_col = _find_col(df, ["x", "easting", "x_coord", "x (easting)"], contains=False) \
        or _find_col(df, ["easting"], contains=True)
    y_col = _find_col(df, ["y", "northing", "y_coord", "y (northing)"], contains=False) \
        or _find_col(df, ["northing"], contains=True)
    ptal_col = _find_col(df, ["ptal"], contains=True)
    ai_col = _find_col(df, ["ai", "access index", "accessindex"], contains=False) \
        or _find_col(df, ["access index", "ai20"], contains=True)
    if x_col is None or y_col is None or ptal_col is None:
        _warn(key, f"could not detect X/Y/PTAL columns in {path.name} "
                   f"(columns: {list(df.columns)[:12]}...)")
        _note(key, "unrecognised PTAL CSV columns")
        return {}
    df = df.copy()
    df["_x"] = pd.to_numeric(df[x_col], errors="coerce")
    df["_y"] = pd.to_numeric(df[y_col], errors="coerce")
    df = df[df["_x"].notna() & df["_y"].notna()]
    gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df["_x"], df["_y"]),
                           crs=27700)
    # Each grid point represents a 100 m cell: a square 50 m buffer in BNG.
    gdf["geometry"] = gdf.geometry.buffer(50, cap_style=3)
    rows = _emit(gdf, key, name_col=ptal_col, want="polygon", assume_epsg=27700,
                 simplify=False,  # they're already minimal 100 m squares
                 props_fn=lambda f: {"ptal": _cell(f, ptal_col),
                                     "ai": _num(f, ai_col)})
    _note(key, how, len(rows))
    return {key: rows}


def _parse_kv(voltage):
    """Max voltage in kilovolts from an OSM voltage tag ('400000;275000'),
    or None."""
    best = None
    for part in str(voltage or "").split(";"):
        part = part.strip().replace(",", "")
        try:
            v = float(part)
        except ValueError:
            continue
        best = v if best is None else max(best, v)
    if best is None:
        return None
    return round(best / 1000.0, 3)


def _closed_line_to_polygon(geom):
    """Substation outlines exported as closed LineStrings -> Polygon; open
    lines -> centroid point (still a usable marker)."""
    try:
        if geom.geom_type == "LineString" and geom.is_ring:
            return Polygon(geom.coords)
    except Exception:
        pass
    return geom.centroid


def build_power_group():
    """power_line + power_substation from the workflow's osmium-extracted
    GeoJSON. Also fills the substation-name cache used by tec_register."""
    src = os.environ.get("OSM_POWER_GEOJSON", "").strip() \
        or os.environ.get("POWER_LINE_SRC", "").strip() \
        or str(RAW / "osm_power.geojson")
    path = Path(src)
    if not path.exists():
        for k in ("power_line", "power_substation"):
            _note(k, "osm_power.geojson missing")
        _warn("power_line/power_substation",
              f"{path} not found. The GitHub workflow extracts it from the UK"
              " OSM PBF (env OSM_PBF_SRC) with osmium-tool; locally, run the"
              " osmium tags-filter/export steps from"
              " .github/workflows/load-datasets.yml or set OSM_POWER_GEOJSON.")
        return {}
    print(f"  [power] reading {path.name} ...")
    gdf = gpd.read_file(path)
    power_col = _find_col(gdf, ["power"])
    if power_col is None:
        _warn("power_line/power_substation", "no 'power' attribute in the "
              "extracted GeoJSON — re-run osmium export keeping the power tag")
        for k in ("power_line", "power_substation"):
            _note(k, "no power tag in extract")
        return {}
    pw = gdf[power_col].astype(str).str.strip().str.lower()
    volt_col = _find_col(gdf, ["voltage"])
    op_col = _find_col(gdf, ["operator"])
    cables_col = _find_col(gdf, ["cables"])
    sub_col = _find_col(gdf, ["substation"])
    name_col = _find_col(gdf, ["name"])
    out = {}

    # power=line only — power=minor_line is deliberately skipped.
    lines = gdf[(pw == "line")
                & gdf.geometry.geom_type.isin(["LineString", "MultiLineString"])]
    out["power_line"] = _emit(
        lines, "power_line", name_col=name_col, want="line",
        props_fn=lambda f: {"voltage": _cell(f, volt_col),
                            "kv": _parse_kv(_cell(f, volt_col)),
                            "operator": _cell(f, op_col),
                            "cables": _cell(f, cables_col)})
    _note("power_line", "osmium extract", len(out["power_line"]))

    subs = gdf[pw == "substation"].copy()
    if not subs.empty:
        gt = subs.geometry.geom_type
        fix = ~gt.isin(["Polygon", "MultiPolygon", "Point", "MultiPoint"])
        if fix.any():
            subs.loc[fix, "geometry"] = subs.loc[fix].geometry.apply(
                _closed_line_to_polygon)
        # The frontend renders substations as a CIRCLE layer, which draws
        # Points only — a polygon footprint would be invisible. Collapse every
        # non-point to a representative interior point.
        gt = subs.geometry.geom_type
        poly = ~gt.isin(["Point"])
        if poly.any():
            subs.loc[poly, "geometry"] = subs.loc[poly].geometry.representative_point()
    out["power_substation"] = _emit(
        subs, "power_substation", name_col=name_col, want="any",
        props_fn=lambda f: {"voltage": _cell(f, volt_col),
                            "kv": _parse_kv(_cell(f, volt_col)),
                            "operator": _cell(f, op_col),
                            "substation": _cell(f, sub_col)})
    _note("power_substation", "osmium extract", len(out["power_substation"]))

    # Cache (name, centroid) for the tec_register fuzzy join.
    if not subs.empty:
        subs4326 = _ensure_4326(subs)
        for _, f in subs4326.iterrows():
            nm = _cell(f, name_col)
            if nm and f.geometry is not None and not f.geometry.is_empty:
                _SUBSTATIONS.append((nm, f.geometry.centroid))
    print(f"  [power] cached {len(_SUBSTATIONS)} named substations for "
          "tec_register geocoding")
    return out


def _maybe_unzip_geo(path, key, prefer=()):
    """If `path` is actually a ZIP (checked by magic bytes, not extension —
    downloads are cached under a fixed name), extract it and return the best
    vector member. Ranking: 1) a member whose name contains one of the
    `prefer` tokens (e.g. the EA CAMS zip ships six variants and we want the
    Q95 low-flow one); 2) a name suggesting WGS84/4326 (NESO ships BNG and
    WGS84 copies); 3) any member whose first coordinate looks like lon/lat
    (|x| <= 360); 4) the first vector file. Non-zip paths pass through."""
    import zipfile
    try:
        with open(path, "rb") as fh:
            if fh.read(4) != b"PK\x03\x04":
                return path
    except OSError:
        return path
    dest = RAW / f"{key}_unzipped"
    dest.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path) as zf:
        zf.extractall(dest)
    vecs = sorted([f for f in dest.rglob("*")
                   if f.suffix.lower() in (".geojson", ".json", ".gpkg", ".shp")])
    if not vecs:
        _warn(key, f"zip source {path.name} holds no vector file")
        return None
    for tok in prefer:
        hits = [f for f in vecs if tok.lower() in f.name.lower()]
        if hits:
            print(f"  [{key}] zip source: using preferred member {hits[0].name}")
            return hits[0]
    named = [f for f in vecs if any(t in f.name.lower()
                                    for t in ("wgs84", "wgs_84", "4326"))]
    if named:
        print(f"  [{key}] zip source: using WGS84 member {named[0].name}")
        return named[0]
    for f in vecs:
        if f.suffix.lower() in (".geojson", ".json"):
            # Cheap probe: the first coordinate pair's magnitude tells BNG
            # (six-figure metres) apart from lon/lat without a full parse.
            head = f.read_text(encoding="utf-8-sig", errors="ignore")[:100000]
            m = re.search(r"\[\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)", head)
            if m and abs(float(m.group(1))) <= 360 and abs(float(m.group(2))) <= 360:
                print(f"  [{key}] zip source: {f.name} looks like lon/lat")
                return f
    print(f"  [{key}] zip source: falling back to {vecs[0].name}")
    return vecs[0]


def build_gsp_boundary():
    key = "gsp_boundary"
    path, how = _resolve_source(key, ["GSP_BOUNDARY_SRC", "GSP_SRC"],
                                GSP_DEFAULT_URL, "gsp-boundaries.geojson")
    if path is not None:
        path = _maybe_unzip_geo(path, key)
    if path is None:
        _warn(key, f"{how} — the NESO resource hash changes between releases."
                   " Set GSP_SRC to the current 'GIS Boundaries for GB Grid"
                   " Supply Points' GeoJSON URL from the NESO Data Portal"
                   " (https://www.neso.energy/data-portal).")
        _note(key, how)
        return {}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    gdf = gpd.read_file(path)
    name_col = _find_col(gdf, ["gsp", "gsp_name", "gspname", "name", "gsps",
                               "sitename"], contains=True)
    id_cols = [c for c in gdf.columns
               if c != "geometry" and re.search(
                   r"(^|_)(id|code|gsp|group)($|_)", str(c).lower())]

    def props_fn(f, cols=tuple(id_cols)):
        return {_slug_key(c): _cell(f, c) for c in cols}

    rows = _emit(gdf, key, name_col=name_col, want="polygon", props_fn=props_fn)
    _note(key, how, len(rows))
    return {key: rows}


def build_ukpn_sites():
    """UKPN grid & primary substations with headroom attributes. All source
    columns are kept as props (names vary; the click card shows everything)."""
    key = "ukpn_sites"
    path, how = _resolve_source(key, ["UKPN_SITES_SRC"], UKPN_SITES_URL,
                                "ukpn-sites.geojson")
    if path is None:
        _warn(key, f"{how} — if the anonymous export is blocked, log into "
                   "ukpowernetworks.opendatasoft.com and set UKPN_SITES_SRC "
                   "to an authorised export URL.")
        _note(key, how)
        return {}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    gdf = gpd.read_file(path)
    name_col = _find_col(gdf, ["sitename", "site_name", "site", "name"],
                         contains=True)
    cols = [c for c in gdf.columns if c != "geometry"][:40]

    def props_fn(f):
        out = {}
        for c in cols:
            v = _jsonable(f[c])
            if v not in (None, ""):
                out[_slug_key(c)] = v
        return out

    rows = _emit(gdf, key, name_col=name_col, want="any", props_fn=props_fn)
    _note(key, how, len(rows))
    return {key: rows}


def build_nged_sites():
    """NGED substation capacity CSV -> points. Column names vary by release;
    detect coordinates defensively (lat/long or easting/northing) and keep all
    other columns as props for the click card."""
    key = "nged_sites"
    path, how = _resolve_source(key, ["NGED_SITES_SRC"], NGED_SITES_URL,
                                "nged-sites.csv")
    if path is None:
        _warn(key, f"{how} — set NGED_SITES_SRC to the substations CSV from "
                   "connecteddata.nationalgrid.co.uk")
        _note(key, how)
        return {}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    df = _read_csv(path)
    # NGED's export lists every SECONDARY (street-level 11 kV) substation —
    # ~120k rows of noise for capacity mapping. Keep the Primary/BSP/GSP tiers,
    # where the connection-capacity story lives (mirrors UKPN's granularity).
    type_col = _find_col(df, ["type"], contains=False)
    if type_col is not None:
        before = len(df)
        df = df[~df[type_col].astype(str).str.strip().str.lower().eq("secondary")]
        if before - len(df):
            print(f"  [{key}] dropped {before - len(df):,} secondary substations "
                  f"({len(df):,} Primary/BSP/GSP kept)")
    lat_col = _find_col(df, ["latitude", "lat"], contains=True)
    lon_col = _find_col(df, ["longitude", "long", "lng"], contains=True)
    east_col = _find_col(df, ["easting", "x"], contains=True)
    north_col = _find_col(df, ["northing", "y"], contains=True)
    name_col = _find_col(df, ["substation name", "site name", "name",
                              "substation"], contains=True)
    df = df.copy()
    if lat_col is not None and lon_col is not None:
        df["_lat"] = pd.to_numeric(df[lat_col], errors="coerce")
        df["_lon"] = pd.to_numeric(df[lon_col], errors="coerce")
        df = df[df["_lat"].notna() & df["_lon"].notna()]
        gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df["_lon"], df["_lat"]),
                               crs=4326)
    elif east_col is not None and north_col is not None:
        df["_e"] = pd.to_numeric(df[east_col], errors="coerce")
        df["_n"] = pd.to_numeric(df[north_col], errors="coerce")
        df = df[df["_e"].notna() & df["_n"].notna()]
        gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df["_e"], df["_n"]),
                               crs=27700).to_crs(4326)
    else:
        _warn(key, f"no coordinate columns detected in {path.name} "
                   f"(columns: {list(df.columns)[:12]}...)")
        _note(key, "no coordinate columns in CSV")
        return {}
    cols = [c for c in gdf.columns
            if c not in ("geometry", "_lat", "_lon", "_e", "_n")][:40]

    def props_fn(f):
        out = {}
        for c in cols:
            v = _jsonable(f[c])
            if v not in (None, ""):
                out[_slug_key(c)] = v
        return out

    rows = _emit(gdf, key, name_col=name_col, want="point", props_fn=props_fn)
    _note(key, how, len(rows))
    return {key: rows}


def _dno_points(df, key):
    """CSV -> point GeoDataFrame via defensive coordinate sniffing (lat/lon,
    easting/northing, or a WKT/`lat, lon` combined column). Returns None with
    a warning when nothing coordinate-like exists."""
    lat_col = _find_col(df, ["latitude", "lat"], contains=True)
    lon_col = _find_col(df, ["longitude", "long", "lng"], contains=True)
    east_col = _find_col(df, ["easting", "x"], contains=True)
    north_col = _find_col(df, ["northing", "y"], contains=True)
    geo_col = _find_col(df, ["geo_point", "geopoint", "location", "spatial",
                             "coordinates"], contains=True)
    df = df.copy()
    if lat_col is not None and lon_col is not None:
        df["_lat"] = pd.to_numeric(df[lat_col], errors="coerce")
        df["_lon"] = pd.to_numeric(df[lon_col], errors="coerce")
    elif east_col is not None and north_col is not None:
        df["_e"] = pd.to_numeric(df[east_col], errors="coerce")
        df["_n"] = pd.to_numeric(df[north_col], errors="coerce")
        df = df[df["_e"].notna() & df["_n"].notna()]
        return gpd.GeoDataFrame(
            df, geometry=gpd.points_from_xy(df["_e"], df["_n"]),
            crs=27700).to_crs(4326)
    elif geo_col is not None:
        # opendatasoft "geo_point_2d" style: "lat, lon" in one cell.
        parts = df[geo_col].astype(str).str.extract(
            r"(-?\d+\.?\d*)[,;\s]+(-?\d+\.?\d*)")
        df["_lat"] = pd.to_numeric(parts[0], errors="coerce")
        df["_lon"] = pd.to_numeric(parts[1], errors="coerce")
    else:
        _warn(key, f"no coordinate columns detected "
                   f"(columns: {list(df.columns)[:12]}...)")
        return None
    df = df[df["_lat"].notna() & df["_lon"].notna()]
    # Sanity: swapped lat/lon (portals disagree) — UK lat is 49..61.
    if len(df) and not df["_lat"].between(49, 61.5).mean() > 0.5:
        df[["_lat", "_lon"]] = df[["_lon", "_lat"]]
        df = df[df["_lat"].between(49, 61.5)]
    return gpd.GeoDataFrame(
        df, geometry=gpd.points_from_xy(df["_lon"], df["_lat"]), crs=4326)


def build_dno_group():
    """The four remaining DNOs (SPEN / NPG / ENWL / SSEN) as *_sites point
    datasets, mirroring ukpn_sites/nged_sites: every source column kept as a
    prop so the click card shows the operator's own headroom vocabulary."""
    out = {}
    for key, env, default_url, drop_name, label in DNO_SITES:
        # "|"-separated URLs (env or default) concatenate into one layer —
        # operators like SPEN split their network across licence areas.
        srcs = os.environ.get(env, "").strip() or (default_url or "")
        parts = [s.strip() for s in srcs.split("|") if s.strip()]
        stem, ext = (drop_name.rsplit(".", 1) + ["csv"])[:2]
        resolved = []
        if parts:
            for i, u in enumerate(parts):
                cname = drop_name if i == 0 else f"{stem}-{i + 1}.{ext}"
                if u.lower().startswith(("http://", "https://")):
                    path, how = _resolve_source(key, [], u, cname)
                else:
                    p = Path(u)
                    path, how = (p, "local file") if p.exists() \
                        else (None, f"{u} does not exist")
                if path is None:
                    _warn(key, f"{how} — set {env} to a CSV export from the "
                               f"{label} open data portal "
                               "(docs/MANUAL_TASKS.md 9b)")
                else:
                    resolved.append((path, how))
        else:
            path, how = _resolve_source(key, [env], None, drop_name)
            if path is not None:
                resolved.append((path, how))
            else:
                _warn(key, f"{how} — set {env} to a CSV export from the "
                           f"{label} open data portal (docs/MANUAL_TASKS.md 9b)")
        if not resolved:
            _note(key, "no reachable source")
            continue
        dfs = []
        for path, how in resolved:
            print(f"  [{key}] reading {path.name} ({how}) ...")
            try:
                dfs.append(_read_csv(path))
            except Exception as e:
                _warn(key, f"unreadable CSV {path.name} ({e}) — check {env}")
        if not dfs:
            _note(key, "unreadable CSV")
            continue
        df = pd.concat(dfs, ignore_index=True, sort=False) \
            if len(dfs) > 1 else dfs[0]
        gdf = _dno_points(df, key)
        if gdf is None or not len(gdf):
            _note(key, "no mappable rows")
            continue
        name_col = _find_col(gdf, ["substation name", "site name", "sitename",
                                   "site", "name"], contains=True)
        cols = [c for c in gdf.columns
                if c not in ("geometry", "_lat", "_lon", "_e", "_n")][:40]

        def props_fn(f, _cols=cols):
            out_p = {}
            for c in _cols:
                v = _jsonable(f[c])
                if v not in (None, ""):
                    out_p[_slug_key(c)] = v
            return out_p

        rows = _emit(gdf, key, name_col=name_col, want="point",
                     props_fn=props_fn)
        _note(key, how, len(rows))
        out[key] = rows
    return out


def _norm_site(s):
    """Normalise a substation/connection-site name for fuzzy matching."""
    s = str(s or "").lower()
    for w in ("grid supply point", "supply point", "substation", "switching",
              "station", "s/s", "gsp", "bsp", "grid", "no.", "number"):
        s = s.replace(w, " ")
    s = re.sub(r"\b\d+\s*k?v\b", " ", s)     # 400kV / 132 kv etc.
    return re.sub(r"[^a-z0-9]+", "", s)


def build_tec_register():
    key = "tec_register"
    path, how = _resolve_source(key, ["TEC_REGISTER_SRC", "TEC_SRC"], None,
                                "tec-register.csv")
    if path is None:
        _warn(key, "no TEC source — set TEC_SRC to a CSV export of the NESO"
                   " TEC Register (https://www.neso.energy/data-portal,"
                   " 'Transmission Entry Capacity (TEC) Register').")
        _note(key, "TEC_SRC not set (no stable default URL)")
        return {}
    if not _SUBSTATIONS:
        _warn(key, "no power_substation features were built this run, so TEC"
                   " rows cannot be geocoded — build power_substation in the"
                   " same run (it needs the osmium OSM extract).")
        _note(key, "no substations available to geocode against")
        return {}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    df = _read_csv(path)
    site_col = _find_col(df, ["connection site", "connection_site", "site name",
                              "connection point", "substation"], contains=True)
    # Cumulative Total Capacity is the CONTRACTED MW (what the queue is worth
    # once built); "MW Connected" is today's figure and reads 0 for every
    # pipeline project. Look it up ALONE first: _find_col's exact-name pass
    # covers all candidates before its contains pass, so "MW Connected" (an
    # exact header) would otherwise beat the cumulative column (whose header
    # carries a "(MW)" suffix and only contains-matches).
    mw_col = _find_col(df, ["cumulative total capacity"], contains=True) \
        or _find_col(df, ["capacity (mw)", "mw increase", "mw connected", "mw"],
                     contains=True)
    conn_col = _find_col(df, ["mw connected"], contains=True)
    status_col = _find_col(df, ["project status", "status", "agreement type",
                                "stage"], contains=True)
    cust_col = _find_col(df, ["customer name", "customer", "company",
                              "project name", "developer"], contains=True)
    plant_col = _find_col(df, ["plant type", "technology"], contains=True)
    host_col = _find_col(df, ["host to", "host"], contains=True)
    date_col = _find_col(df, ["mw effective", "effective from",
                              "completion date"], contains=True)
    if site_col is None:
        _warn(key, f"no connection-site column detected in {path.name} "
                   f"(columns: {list(df.columns)[:12]}...)")
        _note(key, "unrecognised TEC CSV columns")
        return {}

    # Normalised substation lookup, longest names first so the most specific
    # substation wins a contains-match.
    subs = [(_norm_site(nm), nm, pt) for nm, pt in _SUBSTATIONS]
    subs = sorted([s for s in subs if len(s[0]) >= 4],
                  key=lambda s: len(s[0]), reverse=True)

    rows, unmatched, seen = [], 0, {}
    for _, r in df.iterrows():
        site = _cell(r, site_col)
        norm = _norm_site(site)
        match = None
        if len(norm) >= 4:
            for snorm, sname, spt in subs:
                if snorm in norm or norm in snorm:
                    match = (sname, spt)
                    break
        if match is None:
            unmatched += 1
            continue
        sname, spt = match
        ewkt = "SRID=4326;" + to_wkt(spt, rounding_precision=COORD_DP, trim=True)
        sid = _stable_id(key, f"{site}|{_cell(r, cust_col)}|{ewkt}")
        seen[sid] = seen.get(sid, 0) + 1
        if seen[sid] > 1:
            sid = f"{sid}-{seen[sid]}"
        props = {k: v for k, v in {
            "mw": _num(r, mw_col),
            "mw_connected": _num(r, conn_col),
            "status": _cell(r, status_col),
            "customer": _cell(r, cust_col),
            "plant": _cell(r, plant_col),
            "host_to": _cell(r, host_col),
            "effective": _cell(r, date_col),
            "site": site,
            "substation": sname,
        }.items() if v not in (None, "")}
        rows.append({"dataset": key, "source_id": sid,
                     "name": _cell(r, cust_col) or site or None,
                     "props": props, "geom_wkt": ewkt})
    total = len(rows) + unmatched
    rate = (100.0 * len(rows) / total) if total else 0.0
    print(f"  [{key}] matched {len(rows)}/{total} rows ({rate:.0f}%) to "
          "substation names")
    if unmatched:
        _warn(key, f"{unmatched} TEC row(s) had no substation name match and "
                   "were dropped")
    _note(key, f"{how}; match rate {rate:.0f}%", len(rows))
    return {key: rows}


def build_rents_group():
    """lad_boundary (always, when the boundaries fetch works) and la_rents
    (when an ONS PIPR rents CSV is supplied and joins by LAD code)."""
    out = {}
    bpath, bhow = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn("lad_boundary", f"{bhow} — set LAD_BOUNDARIES_SRC to an ONS Open"
              " Geography LAD boundaries GeoJSON/FeatureServer query URL")
        _note("lad_boundary", bhow)
        _note("la_rents", "no LAD boundaries to join rents onto")
        return out
    print(f"  [lad_boundary] reading {bpath.name} ({bhow}) ...")
    gdf = gpd.read_file(bpath)
    code_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*cd", str(c).lower()):
            code_col = c
            break
    code_col = code_col or _find_col(gdf, ["lad_code", "code", "areacd",
                                           "area_code"], contains=True)
    name_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*nm", str(c).lower()):
            name_col = c
            break
    name_col = name_col or _find_col(gdf, NAME_CANDIDATES)

    out["lad_boundary"] = _emit(
        gdf, "lad_boundary", name_col=name_col, id_col=code_col, want="polygon",
        props_fn=lambda f: {"lad_code": _cell(f, code_col)})
    _note("lad_boundary", bhow, len(out["lad_boundary"]))

    rpath, rhow = _resolve_source("la_rents", ["LA_RENTS_SRC", "ONS_RENTS_SRC"],
                                  None, "ons-la-rents.csv")
    if rpath is None:
        _warn("la_rents", "no rents CSV — set ONS_RENTS_SRC to the latest ONS"
              " PIPR (Price Index of Private Rents) local-authority-level CSV"
              " (no stable download URL; grab it from the ONS PIPR release"
              " page). lad_boundary was still built so the boundaries layer"
              " works alone.")
        _note("la_rents", "ONS_RENTS_SRC not set")
        return out
    print(f"  [la_rents] reading {rpath.name} ({rhow}) ...")
    df = _read_csv(rpath)
    rcode_col = _find_col(df, ["area code", "areacd", "ons code", "lad code",
                               "ladcd", "local authority code", "geography code",
                               "code"], contains=True)
    if rcode_col is None:
        _warn("la_rents", f"no LAD-code column detected in {rpath.name} "
              f"(columns: {list(df.columns)[:12]}...)")
        _note("la_rents", "unrecognised rents CSV columns")
        return out

    # Rent value columns, detected defensively. Two shapes are handled:
    #  wide  — one row per LAD, columns like 'Mean rent', 'Lower quartile',
    #          '1 bedroom', ...
    #  long  — many rows per LAD with a bedrooms/category column and a single
    #          rent/price value column.
    bed_cat_col = _find_col(df, ["bedroom category", "bedrooms", "bedroom",
                                 "property type"], contains=True)
    value_cols = [c for c in df.columns if c not in (rcode_col, bed_cat_col)
                  and re.search(r"rent|price|mean|median|quartile|bed",
                                str(c).lower())]
    props_by_code = {}
    if bed_cat_col and value_cols:
        # long format: fold each (code, bedroom-category) row into props.
        val_col = value_cols[0]
        for c in value_cols:
            if re.search(r"rent|price", str(c).lower()):
                val_col = c
                break
        for _, r in df.iterrows():
            code = _cell(r, rcode_col)
            v = _num(r, val_col)
            if not code or v is None:
                continue
            cat = _slug_key(_cell(r, bed_cat_col)) or "all"
            p = props_by_code.setdefault(code, {})
            p[f"rent_{cat}"] = v
            if "all" in cat:
                p.setdefault("rent_mean", v)
    elif value_cols:
        # wide format: one row per LAD.
        for _, r in df.iterrows():
            code = _cell(r, rcode_col)
            if not code:
                continue
            p = props_by_code.setdefault(code, {})
            for c in value_cols:
                v = _num(r, c)
                if v is None:
                    continue
                lc = str(c).lower()
                if "mean" in lc or "average" in lc:
                    p["rent_mean"] = v
                elif "lower quartile" in lc or re.search(r"\blq\b", lc):
                    p["rent_lq"] = v
                else:
                    p[_slug_key(c)] = v
    if not props_by_code:
        _warn("la_rents", "no usable rent values detected in the CSV")
        _note("la_rents", "no rent values detected")
        return out

    joined = gdf[gdf[code_col].astype(str).isin(props_by_code)] if code_col else gdf
    print(f"  [la_rents] {len(joined)}/{len(gdf)} LADs matched a rents row")

    def rent_props(f):
        p = dict(props_by_code.get(_cell(f, code_col), {}))
        p["lad_code"] = _cell(f, code_col)
        return p

    out["la_rents"] = _emit(joined, "la_rents", name_col=name_col,
                            id_col=code_col, want="polygon", props_fn=rent_props)
    _note("la_rents", f"{rhow}; joined on {rcode_col}", len(out["la_rents"]))
    return out


def build_hdt():
    """Housing Delivery Test measurements joined onto LAD polygons — the
    strongest single NPPF approval signal (<75% delivery triggers the
    presumption in favour of sustainable development). Supply the latest
    published HDT measurement CSV as HDT_SRC or a drop-in `hdt.csv`."""
    key = "hdt"
    hpath, hhow = _resolve_source(key, ["HDT_SRC"], None, "hdt.csv")
    if hpath is None:
        _warn(key, "no HDT CSV — download the latest Housing Delivery Test "
                   "measurement from gov.uk (search 'housing delivery test "
                   "measurement') and drop it in as hdt.csv or set HDT_SRC.")
        _note(key, "HDT_SRC not set")
        return {}
    bpath, bhow = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries to join HDT onto")
        _note(key, "no LAD boundaries")
        return {}
    print(f"  [{key}] reading {hpath.name} ({hhow}) ...")
    df = _read_csv(hpath)
    print(f"  [{key}] columns: {list(df.columns)[:10]}")
    code_col = _find_col(df, ["area code", "ons code", "la code", "lpa code",
                              "code"], contains=True)
    meas_col = _find_col(df, ["measurement"], contains=True) \
        or _find_col(df, ["hdt"], contains=True)
    cons_col = _find_col(df, ["consequence"], contains=True)
    name_hcol = _find_col(df, ["authority", "name"], contains=True)
    if meas_col is None:
        _warn(key, f"no measurement column in {hpath.name}")
        _note(key, "no measurement column")
        return {}
    by_code, by_name = {}, {}
    for _, r in df.iterrows():
        m = _num(r, meas_col)
        if m is None:
            continue
        if m <= 2:                      # published as a fraction, not a %
            m = round(m * 100, 1)
        entry = {"hdt_pct": m}
        if cons_col:
            entry["consequence"] = _cell(r, cons_col)
        code = _cell(r, code_col) if code_col else ""
        if code:
            by_code[code.strip()] = entry
        nm = _cell(r, name_hcol) if name_hcol else ""
        if nm:
            by_name[re.sub(r"[^a-z0-9]", "", nm.lower())] = entry
    gdf = gpd.read_file(bpath)
    gcode = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*cd", str(c).lower()):
            gcode = c
            break
    gname = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*nm", str(c).lower()):
            gname = c
            break

    def hdt_props(f):
        e = by_code.get(_cell(f, gcode)) if gcode else None
        if e is None and gname:
            e = by_name.get(re.sub(r"[^a-z0-9]", "", _cell(f, gname).lower()))
        return dict(e) if e else {}

    keep = gdf[gdf.apply(lambda f: bool(hdt_props(f)), axis=1)]
    print(f"  [{key}] {len(keep)}/{len(gdf)} authorities matched an HDT row")
    rows = _emit(keep, key, name_col=gname, id_col=gcode, want="polygon",
                 props_fn=hdt_props)
    _note(key, hhow, len(rows))
    return {key: rows}


def build_building_height():
    """Building heights from OpenStreetMap `height` / `building:levels` tags,
    as points (one per building centroid) classified low / mid / high rise.

    Coverage is whatever OSM contributors have tagged — excellent in city
    centres and for tall/landmark buildings, patchy for suburban housing — so
    the layer is honest about being a sample, not a survey. Single-storey
    sheds and garages are dropped (storeys >= 2 or height >= 6 m) to keep the
    national volume sane; the frontend thins further by height at wide zooms.

    The workflow's osmium step writes data/raw/osm_building_height.geojson.
    """
    key = "building_height"
    path = RAW / "osm_building_height.geojson"
    if not path.exists():
        _warn(key, "no osm_building_height.geojson — run via the workflow "
                   "(its osmium step extracts tagged buildings from the UK PBF)")
        _note(key, "no OSM extract")
        return {}
    print(f"  [{key}] reading {path.name} ...")
    gdf = gpd.read_file(path)
    gdf = gdf[gdf.geometry.notna() & ~gdf.geometry.is_empty].copy()

    def tagval(s, k):
        if not isinstance(s, str):
            return None
        m = re.search('"' + k + '"=>"([^"]*)"', s)
        return m.group(1) if m else None

    def _multi(v):
        """OSM records a way spanning parts of differing height as
        semicolon-separated values ("4;1"). Stripping the punctuation and
        reading the rest as one number gives 41 storeys — see
        pipeline/build_building_tiles.py, which shares these rules."""
        return [p for p in re.split(r"\s*;\s*", str(v).strip()) if p]

    def _parse_one_height(v):
        v = v.strip().lower()
        m = re.match(r"^([\d.]+)\s*(m|metres?|meters?)?$", v)
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                return None
        m = re.match(r"^([\d.]+)\s*(ft|feet|')$", v)
        if m:
            try:
                return float(m.group(1)) * 0.3048
            except ValueError:
                return None
        return None

    def parse_height(v):
        """OSM height: metres, sometimes with units or feet, sometimes several
        values for one way — the tallest part is what matters."""
        if not v:
            return None
        vals = [h for h in (_parse_one_height(p) for p in _multi(v)) if h is not None]
        return max(vals) if vals else None

    ot = gdf["other_tags"] if "other_tags" in gdf.columns else None
    heights, storeys = [], []
    for i in range(len(gdf)):
        tags = ot.iloc[i] if ot is not None else None
        h = parse_height(gdf["height"].iloc[i] if "height" in gdf.columns else None) \
            or parse_height(tagval(tags, "height"))
        lv = None
        raw_lv = tagval(tags, "building:levels")
        if raw_lv:
            parts = []
            for part in _multi(raw_lv):
                try:
                    n = float(re.sub(r"[^\d.]", "", part) or 0)
                except ValueError:
                    continue
                if n:
                    parts.append(n)
            lv = max(parts) if parts else None
        # Either measure can stand in for the other: ~3.2 m per storey.
        if h is None and lv:
            h = round(lv * 3.2, 1)
        if lv is None and h:
            lv = round(h / 3.2)
        heights.append(h)
        storeys.append(lv)
    gdf["_h"] = heights
    gdf["_lv"] = storeys
    before = len(gdf)
    gdf = gdf[gdf["_h"].notna()]
    gdf = gdf[(gdf["_h"] >= 6) | (gdf["_lv"].fillna(0) >= 2)]
    # Sanity: drop obvious tagging errors (a 900 m house).
    gdf = gdf[(gdf["_h"] > 1.5) & (gdf["_h"] < 400)]
    print(f"  [{key}] {before:,} tagged buildings -> {len(gdf):,} usable "
          "(>=2 storeys or >=6 m)")
    if not len(gdf):
        _note(key, "no usable heights")
        return {}
    # Points keep a national layer tractable; footprints stay in OSM.
    gdf["geometry"] = gdf.geometry.centroid
    name_col = _find_col(gdf, ["name"], contains=False)

    def hclass(h):
        if h >= 25:
            return "high"
        if h >= 12:
            return "mid"
        return "low"

    def props_fn(f):
        h = float(f["_h"])
        out = {"height_m": round(h, 1), "class": hclass(h)}
        if f["_lv"]:
            out["storeys"] = int(f["_lv"])
        return out

    rows = _emit(gdf, key, name_col=name_col, want="point", simplify=False,
                 props_fn=props_fn)
    _note(key, "OSM height/building:levels", len(rows))
    return {key: rows}


def build_ofcom_fibre():
    """Ofcom Connected Nations fixed-coverage stats joined onto LAD polygons
    — full-fibre / gigabit availability per authority, the data-centre and
    connectivity signal. Ofcom's download links are per-release, so this
    needs OFCOM_FIBRE_SRC (or a drop-in ofcom-fixed-la.csv): the local
    authority level 'fixed coverage' CSV from the latest Connected Nations
    report."""
    key = "ofcom_fibre"
    path, how = _resolve_source(key, ["OFCOM_FIBRE_SRC"], None,
                                "ofcom-fixed-la.csv")
    if path is None:
        _warn(key, "no Ofcom coverage CSV — set OFCOM_FIBRE_SRC to the "
                   "Connected Nations fixed-coverage local-authority CSV "
                   "(docs/MANUAL_TASKS.md 11)")
        _note(key, how)
        return {}
    bpath, _b = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries to join coverage onto")
        _note(key, "no LAD boundaries")
        return {}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    df = _read_csv(path)
    print(f"  [{key}] columns: {list(df.columns)[:12]}")
    code_col = _find_col(df, ["laua", "la code", "lacode", "area code",
                              "ons code", "code"], contains=True)
    fttp_col = _find_col(df, ["full fibre availability", "fttp availability",
                              "fttp"], contains=True)
    gig_col = _find_col(df, ["gigabit availability", "gigabit"],
                        contains=True)
    if code_col is None or (fttp_col is None and gig_col is None):
        _warn(key, "couldn't find LA-code + FTTP/gigabit columns — is this "
                   "the LA-level fixed coverage file?")
        _note(key, "unrecognised columns")
        return {}
    by_code = {}
    for _, r in df.iterrows():
        code = _cell(r, code_col)
        if not code:
            continue
        e = {}
        for col, name in ((fttp_col, "fttp_pct"), (gig_col, "gigabit_pct")):
            if col is None:
                continue
            v = _num(r, col)
            if v is None:
                continue
            if v <= 1.5:               # published as a fraction
                v *= 100
            e[name] = round(v, 1)
        if e:
            by_code[code.strip()] = e
    gdf = gpd.read_file(bpath)
    gcode = next((c for c in gdf.columns
                  if re.fullmatch(r"lad\d*cd", str(c).lower())), None)
    gname = next((c for c in gdf.columns
                  if re.fullmatch(r"lad\d*nm", str(c).lower())), None)

    def fibre_props(f):
        e = by_code.get(_cell(f, gcode)) if gcode else None
        return dict(e) if e else {}

    keep = gdf[gdf.apply(lambda f: bool(fibre_props(f)), axis=1)]
    print(f"  [{key}] {len(keep)}/{len(gdf)} authorities matched")
    rows = _emit(keep, key, name_col=gname, id_col=gcode, want="polygon",
                 props_fn=fibre_props)
    _note(key, how, len(rows))
    return {key: rows}


# National-average new-build base costs, £/m² GIA, used with the regional
# factor to seed the viability model's typology costs. A PROXY, not BCIS
# (commercial, licensed — excluded by project policy): mid-range figures from
# openly published new-build cost commentary, stated to the user as such and
# overridable both here (env) and per project in the Viability variables modal.
BUILD_COST_HOUSE_PM2 = float(os.environ.get("BUILD_COST_HOUSE_PM2") or 1650)
BUILD_COST_FLAT_PM2 = float(os.environ.get("BUILD_COST_FLAT_PM2") or 2100)


def build_build_cost():
    """Localised build-cost index on LAD polygons — dataset build_cost_index.

    The factor CSV is COMMITTED (pipeline/data/build_cost_index.csv): one row
    per region, UK = 1.00, every row carrying its source string and as-of date.
    BUILD_COST_SRC (URL or path) overrides it — e.g. a client with a real BCIS
    licence exporting their own location factors keeps the same column shape.

    LAD -> region comes from the committed rents CSV (ons-la-rents.csv), which
    already maps every English/Welsh/Scottish LAD code to its region — no new
    download, and one fewer thing to rot. A LAD absent from that lookup gets
    factor 1.0 with src='default' rather than dropping off the map: a missing
    row would read as "no data", when what we mean is "national average"."""
    key = "build_cost_index"
    bpath, bhow = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries to join cost factors onto")
        _note(key, "no LAD boundaries")
        return {}
    fpath, fhow = _resolve_source(
        key, ["BUILD_COST_SRC"], None, "build-cost-factors.csv")
    if fpath is None:
        committed = ROOT / "pipeline" / "data" / "build_cost_index.csv"
        if committed.exists():
            fpath, fhow = committed, "committed proxy table"
        else:
            _warn(key, "no factor CSV (committed file missing and "
                       "BUILD_COST_SRC unset)")
            _note(key, "no factor table")
            return {}
    print(f"  [{key}] reading {fpath.name} ({fhow}) ...")
    fdf = _read_csv(fpath)
    rcol = _find_col(fdf, ["region"], contains=True)
    fcol = _find_col(fdf, ["factor"], contains=True)
    acol = _find_col(fdf, ["asof", "as of", "date"], contains=True)
    if rcol is None or fcol is None:
        _warn(key, f"factor CSV needs region+factor columns, got "
                   f"{list(fdf.columns)}")
        _note(key, "unrecognised factor CSV")
        return {}
    factors = {}
    for _, row in fdf.iterrows():
        region = str(row[rcol]).strip()
        f = _num(row, fcol)
        if region and f:
            factors[region.lower()] = (f, str(row[acol]).strip() if acol else "")

    # LAD code -> region, from the committed rents lookup.
    lad_region = {}
    rents = ROOT / "data" / "raw" / "ons-la-rents.csv"
    if rents.exists():
        rdf = _read_csv(rents)
        ccol = _find_col(rdf, ["area code", "code"], contains=True)
        gcol = _find_col(rdf, ["region"], contains=True)
        if ccol is not None and gcol is not None:
            for _, row in rdf.iterrows():
                lad_region[str(row[ccol]).strip()] = str(row[gcol]).strip()
    if not lad_region:
        _warn(key, "LAD->region lookup unavailable (ons-la-rents.csv) — "
                   "every authority will carry the national factor 1.0")

    gdf = gpd.read_file(bpath)
    code_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*cd", str(c).lower()):
            code_col = c
            break
    code_col = code_col or _find_col(gdf, ["lad_code", "code", "areacd"],
                                     contains=True)
    name_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*nm", str(c).lower()):
            name_col = c
            break

    n_default = 0

    def cost_props(f):
        nonlocal n_default
        code = str(_cell(f, code_col) or "").strip()
        region = lad_region.get(code, "")
        hit = factors.get(region.lower()) if region else None
        if hit is None:
            n_default += 1
            factor, asof, src = 1.0, "", "default"
        else:
            factor, asof = hit
            src = "proxy"
        return {"lad_code": code, "region": region or None,
                "factor": round(factor, 3),
                "cost_house_pm2": round(BUILD_COST_HOUSE_PM2 * factor),
                "cost_flat_pm2": round(BUILD_COST_FLAT_PM2 * factor),
                "asof": asof or None, "src": src}

    rows = _emit(gdf, key, name_col=name_col, id_col=code_col, want="polygon",
                 props_fn=cost_props)
    if n_default:
        print(f"  [{key}] {n_default} authorit(ies) had no region mapping — "
              f"carrying national factor 1.0, flagged src='default'")
    _note(key, f"{fhow}; base house £{BUILD_COST_HOUSE_PM2:.0f}/m² "
               f"flat £{BUILD_COST_FLAT_PM2:.0f}/m²", len(rows))
    return {key: rows}


LAND_VALUE_DEFAULT_URL = (
    "https://assets.publishing.service.gov.uk/media/"
    "69b2d95f912f3f96bf687b1c/"
    "Land_value_estimates_for_policy_appraisal_2023.xlsx")
LAND_VALUE_ASOF = "2023"


def build_land_value():
    """MHCLG/VOA residential land value estimates on LAD polygons — land_value.

    Source: 'Land value estimates for policy appraisal' (MHCLG, valuations by
    the VOA; OGL) — the value per hectare of a TYPICAL residential site in
    each English local authority. This is the published benchmark the
    viability engine's land line uses in place of the value-ratio proxy where
    available. MHCLG's own caveat travels with the data: these are policy
    appraisal estimates, not market valuations of any specific site.

    The workbook layout isn't guaranteed across releases, so parsing is a
    tolerant grid scan rather than fixed cell references: find a header cell
    containing 'resid', read that column downwards, and take the value for
    any row that also carries an ONS area code. LAND_VALUE_SRC (URL or path,
    xlsx or csv) overrides the default download; the parse aborts loudly —
    never silently — if it recognises fewer than 100 authorities or the
    median is implausible."""
    key = "land_value"
    bpath, bhow = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries to join land values onto")
        _note(key, "no LAD boundaries")
        return {}
    fpath, fhow = _resolve_source(key, ["LAND_VALUE_SRC"],
                                  LAND_VALUE_DEFAULT_URL,
                                  "land-value-estimates.xlsx")
    if fpath is None:
        _warn(key, f"land value source unavailable ({fhow})")
        _note(key, "no source")
        return {}
    print(f"  [{key}] reading {fpath.name} ({fhow}) ...")

    code_re = re.compile(r"^[EWS]\d{8}$")

    def _money(v):
        # Bare numeric parse only — plausibility and unit scaling are decided
        # AFTER collection, because the workbook may quote absolute £/ha OR
        # £ millions/ha and only the population of values tells us which.
        try:
            x = float(str(v).replace("£", "").replace(",", "").strip())
        except (TypeError, ValueError):
            return None
        return x if x > 0 else None

    def _norm_name(s):
        s = re.sub(r"[^a-z0-9 ]", " ", str(s).lower())
        s = re.sub(r"\b(city of|county of|royal borough of|london borough of|"
                   r"district|borough|council|ua|the)\b", " ", s)
        return re.sub(r"\s+", " ", s.replace(" and ", " ")).strip()

    import pandas as pd
    if fpath.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        grids = pd.read_excel(fpath, sheet_name=None, header=None)
    else:
        grids = {"csv": pd.read_csv(fpath, header=None, dtype=str,
                                    encoding="utf-8-sig",
                                    on_bad_lines="skip")}

    values, by_name = {}, {}

    # Targeted parse for the actual 2023 workbook layout (seen via the CI
    # diagnostics, run 31016580789): a 'Residential' SHEET — title rows, then
    # a header row containing 'LA Code', then one row per authority whose
    # columns are a matrix of £/ha figures for site-size × density scenarios
    # (plus two-digit dwellings/ha assumption columns). The per-authority
    # benchmark is the MEDIAN of its scenario values — MHCLG's own framing is
    # a 'typical' site, and the median of their published scenarios is the
    # faithful single number. The £50k floor excludes the density columns.
    for sname, g in grids.items():
        s = str(sname).lower()
        if "resid" not in s or "adj" in s:
            continue
        nrow, ncol = g.shape
        hdr = None
        for rr in range(min(nrow, 12)):
            if any("la code" in str(g.iat[rr, cc]).lower() for cc in range(ncol)):
                hdr = rr
                break
        if hdr is None:
            continue
        got = {}
        for rr in range(hdr + 1, nrow):
            row = g.iloc[rr]
            code = next((str(x).strip() for x in row
                         if code_re.match(str(x).strip())), None)
            if not code:
                continue
            nums = sorted(v for v in (_money(x) for x in row)
                          if v is not None and v >= 50_000)
            if nums:
                got[code] = nums[len(nums) // 2]
        if len(got) > len(values):
            values = got

    # Generic fallback (covers a LAND_VALUE_SRC CSV or a future re-layout):
    # find a 'resid'/'land value' header cell, read its column, key rows by
    # ONS code or, failing that, by normalised authority name.
    if len(values) < 100:
        for _sname, g in grids.items():
            nrow, ncol = g.shape
            headers = [(rr, cc) for rr in range(min(nrow, 60))
                       for cc in range(ncol)
                       if "resid" in str(g.iat[rr, cc]).lower()
                       or "land value" in str(g.iat[rr, cc]).lower()]
            for hr, hc in headers:
                got, got_names = {}, {}
                for rr in range(hr + 1, nrow):
                    row = g.iloc[rr]
                    val = _money(g.iat[rr, hc])
                    if val is None:
                        continue
                    code = next((str(x).strip() for x in row
                                 if code_re.match(str(x).strip())), None)
                    if code:
                        got[code] = val
                    nm = next((str(x).strip() for x in row
                               if isinstance(x, str) and len(str(x).strip()) > 3
                               and _money(x) is None
                               and not code_re.match(str(x).strip())), None)
                    if nm:
                        got_names[_norm_name(nm)] = val
                if len(got) > len(values):
                    values = got
                if len(got_names) > len(by_name):
                    by_name = got_names

    def _rescale(d):
        # Values quoted in £ millions (median < 1000) -> absolute £.
        if not d:
            return d
        med0 = sorted(d.values())[len(d) // 2]
        return {k: v * 1e6 for k, v in d.items()} if med0 < 1000 else d

    values, by_name = _rescale(values), _rescale(by_name)

    gdf = gpd.read_file(bpath)
    code_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*cd", str(c).lower()):
            code_col = c
            break
    code_col = code_col or _find_col(gdf, ["lad_code", "code", "areacd"],
                                     contains=True)
    name_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*nm", str(c).lower()):
            name_col = c
            break
    if code_col is None:
        _warn(key, "no LAD code column in the boundary file")
        _note(key, "no boundary code column")
        return {}

    # No usable code rows -> map the name-keyed parse onto boundary codes.
    if len(values) < 100 and by_name and name_col is not None:
        matched = {}
        for _, brow in gdf[[code_col, name_col]].iterrows():
            v = by_name.get(_norm_name(brow[name_col]))
            if v is not None:
                matched[str(brow[code_col]).strip()] = v
        if len(matched) > len(values):
            print(f"  [{key}] no ONS codes in the workbook — matched "
                  f"{len(matched)} authorities by name instead")
            values = matched

    if len(values) < 100:
        # Print enough of the workbook to fix the scan from the CI log alone.
        _warn(key, f"only {len(values)} authorities parsed from {fpath.name} "
                   "— has the workbook layout changed? Set LAND_VALUE_SRC to "
                   "a CSV with an ONS code column and a 'Residential ...' "
                   "value column. Sheet diagnostics follow:")
        for sname, g in grids.items():
            print(f"  [{key}]   sheet '{sname}' shape {g.shape}")
            for rr in range(min(8, g.shape[0])):
                cells = [str(g.iat[rr, cc])[:28] for cc in range(min(8, g.shape[1]))]
                print(f"  [{key}]     r{rr}: {cells}")
        _note(key, "unrecognised source format")
        return {}
    med = sorted(values.values())[len(values) // 2]
    if not (100_000 <= med <= 100_000_000):
        _warn(key, f"median parsed value £{med:,.0f}/ha is implausible — "
                   "refusing to load")
        _note(key, "implausible values")
        return {}

    # England-only publication: emit only matched authorities, so Wales and
    # Scotland render as absence-of-data rather than a misleading zero.
    total = len(gdf)
    gdf = gdf[gdf[code_col].astype(str).str.strip().isin(values.keys())]
    print(f"  [{key}] {len(values)} authorities parsed (median "
          f"£{med:,.0f}/ha); {len(gdf)} of {total} boundaries matched")

    def lv_props(f):
        code = str(_cell(f, code_col) or "").strip()
        return {"lad_code": code,
                "resi_gbp_ha": round(values[code]),
                "asof": LAND_VALUE_ASOF,
                "src": "MHCLG/VOA land value estimates for policy appraisal "
                       "(OGL) — appraisal benchmarks, not site valuations"}

    rows = _emit(gdf, key, name_col=name_col, id_col=code_col, want="polygon",
                 props_fn=lv_props)
    _note(key, f"{fhow}; {len(values)} authorities, median £{med:,.0f}/ha",
          len(rows))
    return {key: rows}


def build_cil_rates():
    """Residential CIL rates by charging authority — dataset cil_rates.

    The rates table is COMMITTED (pipeline/data/cil_rates.csv): one row per
    authority with resi_min/typ/max (£/m², ~2025-indexed) spanning the adopted
    charging schedule's zones, plus the Mayoral CIL2 line for London, and
    explicit status 'none' rows for authorities that never adopted CIL (a real
    £0, distinct from no-data). CIL_RATES_SRC (URL or path) overrides it with
    the same column shape. Joined to LAD polygons BY NORMALISED NAME (the CSV
    is hand-compiled from council documents, which use names not ONS codes).

    Every LAD polygon is emitted so the layer reads completely:
      matched adopted  -> rates + cil_pm2 = resi_typ + mayoral
      matched none     -> cil_pm2 = 0 (genuinely no CIL)
      Scottish LADs    -> cil_pm2 = 0, status no_regime (S75 instead of CIL)
      anything else    -> status unknown, cil_pm2 null (engine falls back to
                          its regional band; renders as no-data, not zero)

    All figures are indicative midpoints of zoned schedules and omit future
    indexation — the viability engine frames them as a floor and lets an exact
    adopted rate override per project."""
    key = "cil_rates"
    bpath, bhow = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries to join CIL rates onto")
        _note(key, "no LAD boundaries")
        return {}
    fpath, fhow = _resolve_source(key, ["CIL_RATES_SRC"], None, "cil-rates.csv")
    if fpath is None:
        committed = ROOT / "pipeline" / "data" / "cil_rates.csv"
        if committed.exists():
            fpath, fhow = committed, "committed rates table"
        else:
            _warn(key, "no CIL rates CSV (committed file missing and "
                       "CIL_RATES_SRC unset)")
            _note(key, "no rates table")
            return {}
    import pandas as pd
    print(f"  [{key}] reading {fpath.name} ({fhow}) ...")
    df = pd.read_csv(fpath, comment="#", dtype=str).fillna("")
    need = {"name", "status", "resi_typ"}
    if not need.issubset(set(df.columns)):
        _warn(key, f"rates CSV needs {sorted(need)} columns, got "
                   f"{list(df.columns)}")
        _note(key, "unrecognised rates CSV")
        return {}

    def _norm_name(s):
        s = re.sub(r"[^a-z0-9 ]", " ", str(s).lower())
        s = re.sub(r"\b(city of|county of|royal borough of|london borough of|"
                   r"district|borough|council|ua|the|city)\b", " ", s)
        return re.sub(r"\s+", " ", s.replace(" and ", " ")).strip()

    by_norm = {}
    for _, r in df.iterrows():
        nm = _norm_name(r["name"])
        if nm:
            by_norm[nm] = r

    gdf = gpd.read_file(bpath)
    code_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*cd", str(c).lower()):
            code_col = c
            break
    code_col = code_col or _find_col(gdf, ["lad_code", "code", "areacd"],
                                     contains=True)
    name_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*nm", str(c).lower()):
            name_col = c
            break

    # ---- Verification pass against the platform's own schedule register ----
    # planning.data.gov.uk publishes community-infrastructure-levy-schedule:
    # one record per charging authority with the adopted date and a link to
    # the schedule document itself. It carries no RATES (those live inside the
    # PDFs), so it cannot replace the compiled table — but it is authoritative
    # on the question the table is most likely to get wrong: WHETHER an
    # authority charges CIL at all. Joined on the LAD code, never by name.
    sched = {}
    n_ver = n_conflict = n_unrated = 0
    spath, _show = _resolve_source(
        key, ["CIL_SCHEDULE_SRC"],
        PLANNING_DATA_BASE + "community-infrastructure-levy-schedule.csv",
        "community-infrastructure-levy-schedule.csv")
    opath, _ohow = _resolve_source(
        key, ["LOCAL_AUTHORITY_SRC"],
        PLANNING_DATA_BASE + "local-authority.csv", "local-authority.csv")
    if spath is not None and opath is not None:
        orgs = {}
        with open(opath, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                orgs[r.get("entity", "")] = r
        with open(spath, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                o = orgs.get(r.get("organisation-entity", ""))
                # County councils are not CIL charging authorities for the
                # districts within them; their records would otherwise attach
                # a schedule to areas that do not charge under it.
                if not o or (o.get("local-authority-type") or "").strip() == "CTY":
                    continue
                lad = (o.get("local-authority-district") or "").strip()
                if not lad:
                    continue
                d = (r.get("adopted-date") or "")[:10]
                prev = sched.get(lad)
                if prev is None or d > prev["adopted"]:
                    sched[lad] = {"adopted": d,
                                  "url": (r.get("document-url") or "").strip(),
                                  "title": (r.get("name") or "").strip()}
        print(f"  [{key}] {len(sched)} charging authorities carry a published "
              f"schedule on the platform")
    else:
        _warn(key, "no CIL schedule register — statuses go unverified")

    SRC_ADOPTED = ("adopted charging schedule, ~2025-indexed INDICATIVE "
                   "midpoints — verify against the council's current schedule")
    n_hit = n_unknown = 0

    def _num(r, k):
        try:
            v = float(str(r.get(k, "")).strip())
            return v if v == v else 0.0   # NaN guard
        except (TypeError, ValueError):
            return 0.0

    def _sched_props(sc):
        """The platform's own evidence, attached verbatim so a user can open
        the schedule and check the number we used."""
        if not sc:
            return {}
        out = {"schedule_url": sc["url"] or None,
               "schedule_name": sc["title"] or None}
        if sc["adopted"]:
            out["schedule_adopted"] = sc["adopted"]
        return out

    def cil_props(f):
        nonlocal n_hit, n_unknown, n_ver, n_conflict, n_unrated
        code = str(_cell(f, code_col) or "").strip()
        nm = str(_cell(f, name_col) or "").strip()
        sc = sched.get(code)
        r = by_norm.get(_norm_name(nm))
        if r is None:
            if code.startswith("S"):
                return {"lad_code": code, "status": "no_regime", "cil_pm2": 0,
                        "asof": "2025",
                        "src": "Scotland has no CIL regime — planning "
                               "obligations run through S75 instead"}
            if sc:
                # We have no rate, but the platform proves a schedule exists.
                # That is worth saying out loud: the engine still falls back to
                # its regional band, and now the popup can hand over the
                # actual document instead of shrugging.
                n_unrated += 1
                return dict({"lad_code": code, "status": "schedule_unrated",
                             "cil_pm2": None,
                             "src": "an adopted charging schedule is published "
                                    "for this authority, but its £/m² rates are "
                                    "not yet compiled — the engine falls back "
                                    "to its regional band"},
                            **_sched_props(sc))
            n_unknown += 1
            return {"lad_code": code, "status": "unknown", "cil_pm2": None,
                    "src": "rate not yet compiled — the viability engine "
                           "falls back to its regional band here"}
        n_hit += 1
        status = str(r["status"]).strip()
        if status == "none":
            # A compiled "no CIL here" against a published schedule is a
            # contradiction, and £0 is the expensive way to be wrong. A DATED
            # schedule is strong enough evidence to withdraw the zero and fall
            # back to the regional band; an undated stub (a bare council link)
            # is not, so that only earns a flag.
            if sc and sc["adopted"]:
                n_conflict += 1
                return dict({"lad_code": code, "status": "schedule_unrated",
                             "cil_pm2": None, "conflict": True,
                             "src": "our table recorded no CIL here, but the "
                                    "platform publishes an adopted charging "
                                    "schedule — the £0 has been withdrawn and "
                                    "the regional band applies until the rate "
                                    "is read off the schedule"},
                            **_sched_props(sc))
            out = {"lad_code": code, "status": "none", "cil_pm2": 0,
                   "asof": str(r.get("asof", "")).strip() or None,
                   "src": "authority has not adopted CIL — the charge is "
                          "genuinely zero (S106 still applies)"}
            if sc:
                n_conflict += 1
                out["conflict"] = True
                out["src"] += " — note the platform holds an undated schedule "\
                              "record for this authority, which is worth checking"
            return dict(out, **_sched_props(sc))
        typ, mn, mx = _num(r, "resi_typ"), _num(r, "resi_min"), _num(r, "resi_max")
        may = _num(r, "mayoral")
        note = str(r.get("note", "")).strip()
        if sc:
            n_ver += 1
        return dict({"lad_code": code, "status": status,
                     "resi_min": round(mn), "resi_typ": round(typ),
                     "resi_max": round(mx), "mayoral": round(may),
                     "cil_pm2": round(typ + may),
                     "asof": str(r.get("asof", "")).strip() or None,
                     "verified": bool(sc),
                     "note": note or None, "src": SRC_ADOPTED},
                    **_sched_props(sc))

    rows = _emit(gdf, key, name_col=name_col, id_col=code_col, want="polygon",
                 props_fn=cil_props)
    print(f"  [{key}] {n_hit} authorities matched from the table, "
          f"{n_unknown} unmatched (regional-band fallback)")
    print(f"  [{key}] verification: {n_ver} compiled rates confirmed by a "
          f"published schedule, {n_unrated} authorities have a schedule but no "
          f"compiled rate, {n_conflict} contradict the table")
    if n_hit < 100:
        _warn(key, f"only {n_hit} authorities matched by name — check the "
                   "CSV's name spellings against the boundary file")
    _note(key, f"{fhow}; {n_hit} authorities matched", len(rows))
    return {key: rows}


def build_lsoa_boundary():
    """LSOA boundaries into the database — dataset lsoa_boundary.

    The polygons already ship with the app for the deprivation tiles
    (web/data/lsoa_imd.geojson, ~33.7k features, committed), but they existed
    ONLY as tiles: the lsoa_imd TABLE holds centroids, so nothing server-side
    could assign a sale to its neighbourhood. rebuild_lsoa_prices() needs real
    polygons to point-in-polygon 2M Land Registry comparables, so they are
    loaded here through the ordinary dataset path.

    Local file, no download: LSOA_BOUNDARY_SRC overrides it if a newer
    vintage is ever dropped in."""
    key = "lsoa_boundary"
    fpath, fhow = _resolve_source(key, ["LSOA_BOUNDARY_SRC"], None,
                                  "lsoa-boundaries.geojson")
    if fpath is None:
        committed = ROOT / "web" / "data" / "lsoa_imd.geojson"
        if committed.exists():
            fpath, fhow = committed, "committed app layer"
        else:
            _warn(key, "no LSOA polygon file (web/data/lsoa_imd.geojson "
                       "missing and LSOA_BOUNDARY_SRC unset)")
            _note(key, "no boundary file")
            return {}
    print(f"  [{key}] reading {fpath.name} ({fhow}) ...")
    gdf = gpd.read_file(fpath)
    code_col = _find_col(gdf, ["lsoa_code", "lsoa11cd", "lsoa21cd", "code"],
                         contains=True)
    lad_col = _find_col(gdf, ["lad_name", "lad_nm", "authority"], contains=True)
    if code_col is None:
        _warn(key, f"no LSOA code column, got {list(gdf.columns)[:12]}")
        _note(key, "no code column")
        return {}

    def lsoa_props(f):
        return {"lsoa_code": str(_cell(f, code_col) or "").strip(),
                "lad_name": (str(_cell(f, lad_col)).strip()
                             if lad_col is not None else None)}

    rows = _emit(gdf, key, name_col=lad_col, id_col=code_col, want="polygon",
                 props_fn=lsoa_props)
    _note(key, f"{fhow}; {len(rows)} LSOA polygons", len(rows))
    return {key: rows}


# Local plan process stages, worst -> best evidence of a settled plan. Used to
# pick which plan speaks for a boundary when several are recorded.
LOCAL_PLAN_STAGE_RANK = {
    "": 0, "draft": 1, "regulation-18": 2, "regulation-19": 3,
    "submitted": 4, "examination": 5, "found-sound": 6, "adopted": 7,
}


def build_local_plan_housing():
    """Local plan housing requirement vs allocation, on plan boundaries —
    dataset local_plan_housing.

    There is NO national dataset of allocated housing SITE polygons: probing
    the planning data platform for site-allocation / housing-allocation /
    local-plan-site slugs returns nothing, and allocations remain a per-council
    publication. What IS national is the plan-level arithmetic, and for finding
    where land is worth promoting it is arguably the more useful half:

      required-housing   what the plan must deliver
      allocated-housing  what it has allocated to sites
      committed-housing  what already has permission
      -> gap             what the authority still has to find

    Three CSVs joined onto the local-plan-boundary polygons we already carry:
      local-plan            plan metadata: name, adopted date, period, process
      local-plan-housing    the housing numbers above
      local-plan-timetable  milestone events (latest = current stage)

    Coverage is honest, not total: ~338 of 364 boundaries have a plan record
    and ~150 publish enough to compute a gap. Boundaries with a plan but no
    usable numbers still carry name/adopted-date/stage and are flagged
    has_numbers=false, so the map distinguishes 'no gap' from 'not
    published' — and a plan that publishes only ONE side of its supply is
    'partial_supply', shown with its components but no gap."""
    key = "local_plan_housing"
    bpath, bhow = _resolve_source(
        "local_plan_boundary", ["LOCAL_PLAN_BOUNDARY_SRC"],
        PLANNING_DATA_BASE + "local-plan-boundary.geojson",
        "local-plan-boundary.geojson")
    if bpath is None:
        _warn(key, "no local plan boundaries to join onto")
        _note(key, "no boundaries")
        return {}

    def _csv(slug, env):
        path, how = _resolve_source(key, [env],
                                    PLANNING_DATA_BASE + f"{slug}.csv",
                                    f"{slug}.csv")
        if path is None:
            return {}, how
        rows = {}
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                rows[r.get("reference", "")] = r
        return rows, how

    plans, phow = _csv("local-plan", "LOCAL_PLAN_SRC")
    if not plans:
        _warn(key, f"no local-plan.csv ({phow})")
        _note(key, "no plan table")
        return {}
    # housing + timetable are keyed by the PLAN reference, not their own
    hous, _ = {}, None
    hpath, _hhow = _resolve_source(key, ["LOCAL_PLAN_HOUSING_SRC"],
                                   PLANNING_DATA_BASE + "local-plan-housing.csv",
                                   "local-plan-housing.csv")
    if hpath is not None:
        with open(hpath, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                hous[r.get("local-plan", "")] = r
    # Milestones come from TWO timetables, because the platform is migrating
    # between them and each is thin on its own:
    #   local-plan-timetable  the older collection — 2,700 rows but only 323
    #                         carry an event-date at all
    #   plan-timetable        the current one (keyed `plan`/`plan-event`, with
    #                         actual-date AND predicted-date) — 8,797 rows,
    #                         3,749 dated
    # Adoption dates matter most: the plan table fills adopted-date for just
    # 104 of 987 plans, and the 5-year-review flag is useless without one. The
    # new timetable adds 253 dated adoptions on top, so the flag goes from
    # mostly silent to mostly answerable. Read old first, new second — later
    # writes win where both have an opinion.
    events, adopted_ev = {}, {}

    def _timetable(env, slug, plan_col, event_col, date_cols):
        path, _ = _resolve_source(key, [env], PLANNING_DATA_BASE + slug, slug)
        if path is None:
            return
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                lp = (r.get(plan_col) or "").strip()
                d = next((v for v in ((r.get(c) or "").strip()[:10]
                                      for c in date_cols) if v), "")
                ev = (r.get(event_col) or "").strip()
                if not (lp and d):
                    continue
                if d > (events.get(lp, ("", ""))[0]):
                    events[lp] = (d, ev)
                if ev in ("plan-adopted", "adopted") and d > adopted_ev.get(lp, ""):
                    adopted_ev[lp] = d

    _timetable("LOCAL_PLAN_TIMETABLE_SRC", "local-plan-timetable.csv",
               "local-plan", "local-plan-event", ("event-date",))
    # predicted-date is deliberately NOT read: a plan the council HOPES to
    # adopt next year has not been adopted, and dating it would age it.
    _timetable("PLAN_TIMETABLE_SRC", "plan-timetable.csv",
               "plan", "plan-event", ("actual-date", "event-date"))

    # boundary reference -> the plan that best speaks for it: prefer one with
    # housing numbers, then the furthest-progressed stage, then most recent.
    by_boundary = {}
    for ref, pl in plans.items():
        b = (pl.get("local-plan-boundary") or "").strip()
        if b:
            by_boundary.setdefault(b, []).append(pl)

    def _int(v):
        try:
            return int(float(str(v).replace(",", "").strip()))
        except (TypeError, ValueError):
            return None

    def _complete(ref):
        """Does this plan's housing row support a trustworthy gap? — a
        requirement AND BOTH core supply components."""
        h = hous.get(ref)
        return bool(h) and all(_int(h.get(k)) is not None for k in
                               ("required-housing", "allocated-housing",
                                "committed-housing"))

    def _plan_key(pl):
        return (1 if _complete(pl["reference"]) else 0,
                1 if pl["reference"] in hous else 0,
                LOCAL_PLAN_STAGE_RANK.get(
                    (pl.get("local-plan-process") or "").strip(), 0),
                (pl.get("adopted-date") or "")[:10],
                (pl.get("period-end-date") or "")[:10])

    today = _dt.date.today()
    n_plan = n_num = 0

    def lph_props(f):
        nonlocal n_plan, n_num
        ref = str(_cell(f, "reference") or "").strip()
        cands = by_boundary.get(ref) or []
        if not cands:
            return {"plan_ref": None, "status": "no_plan_record",
                    "has_numbers": False,
                    "src": "planning.data.gov.uk local-plan (OGL) — no plan "
                           "recorded for this boundary"}
        pl = sorted(cands, key=_plan_key)[-1]
        n_plan += 1
        h = hous.get(pl["reference"], {})
        req = _int(h.get("required-housing")) or _int(pl.get("required-housing"))
        alloc = _int(h.get("allocated-housing"))
        comm = _int(h.get("committed-housing"))
        wind = _int(h.get("windfall-housing"))
        broad = _int(h.get("broad-locations-housing"))
        adopted = (pl.get("adopted-date") or "")[:10] or adopted_ev.get(pl["reference"], "")
        age = None
        if len(adopted) == 10:
            try:
                age = round((today - _dt.date.fromisoformat(adopted)).days / 365.25, 1)
            except ValueError:
                age = None
        ev = events.get(pl["reference"])
        out = {
            "plan_ref": pl["reference"],
            "plan_name": (pl.get("name") or "").strip() or None,
            "adopted": adopted or None,
            "plan_age_yrs": age,
            "period_start": (pl.get("period-start-date") or "")[:10] or None,
            "period_end": (pl.get("period-end-date") or "")[:10] or None,
            "stage": (pl.get("local-plan-process") or "").strip() or None,
            "last_event": ev[1] if ev else None,
            "last_event_date": ev[0] if ev else None,
            "plan_url": (pl.get("documentation-url")
                         or pl.get("document-url") or "").strip() or None,
            "required": req, "allocated": alloc, "committed": comm,
            "windfall": wind, "broad_locations": broad,
            "src": "planning.data.gov.uk local-plan + local-plan-housing "
                   "(OGL v3) — plan-level figures as published by the LPA",
        }
        # The number that matters: what the plan still has to find. It needs
        # the requirement AND BOTH core supply components — allocations and
        # commitments. A blank cell means the LPA published no figure, NOT
        # zero, and treating it as zero is not a conservative assumption, it
        # is a wrong one: it was inventing 30,748 homes of shortfall for
        # Cornwall (no allocations published) and 24,325 for Sheffield (no
        # commitments published), which put exactly the wrong authorities at
        # the top of the ranking. 47 of 190 areas were affected, so a quarter
        # of the map — and the loudest quarter. Those now read as
        # 'partial_supply': the published components are kept and shown, the
        # gap is withheld.
        # Windfall and broad locations ARE counted where published: they are
        # supply components of the plan's own trajectory, not extras.
        if req is not None and alloc is not None and comm is not None:
            supply = alloc + comm + (wind or 0) + (broad or 0)
            out["supply"] = supply
            out["gap"] = max(0, req - supply)
            out["pct_met"] = round(100.0 * supply / req, 1) if req > 0 else None
            out["has_numbers"] = True
            out["status"] = "has_numbers"
            n_num += 1
        else:
            out["has_numbers"] = False
            out["status"] = ("partial_supply"
                             if any(v is not None
                                    for v in (req, alloc, comm, wind, broad))
                             else "plan_only")
            if out["status"] == "partial_supply":
                out["missing"] = ", ".join(
                    n for n, v in (("requirement", req), ("allocations", alloc),
                                   ("commitments", comm)) if v is None)
        return out

    gdf = gpd.read_file(bpath)
    ref_col = _find_col(gdf, ["reference"])
    if ref_col is None:
        _warn(key, "no reference column on the boundary file")
        _note(key, "no boundary reference")
        return {}
    name_col = _find_col(gdf, ["name"])
    rows = _emit(gdf, key, name_col=name_col, id_col=ref_col, want="polygon",
                 props_fn=lph_props)
    print(f"  [{key}] {len(rows)} plan boundaries; {n_plan} with a plan "
          f"record, {n_num} with housing numbers")
    _note(key, f"{phow}; {n_num} areas with requirement vs supply", len(rows))
    return {key: rows}


# MHCLG Live Table 125 (dwelling stock by local authority) and the ONS
# affordability ratio are both spreadsheets on pages without a stable file
# URL, so both take the usual *_SRC override and degrade to a warning.
DWELLING_STOCK_URL = ("https://assets.publishing.service.gov.uk/media/"
                      "6a0dcbf65c3c79da61662e39/LiveTable125.ods")
AFFORDABILITY_URL = ("https://www.ons.gov.uk/file?uri=/peoplepopulationand"
                     "community/housing/datasets/ratioofhousepricetoworkplace"
                     "basedearningslowerquartileandmedian/current/"
                     "aff1ratioofhousepricetoworkplacebasedearnings.xlsx")

# NPPF (Aug 2026) Annex D, the standard method. Both constants are the
# Framework's, not ours — do not "tune" them.
LHN_STOCK_PCT = 0.008        # step 1: baseline is 0.8% of existing stock
LHN_RATIO_FLOOR = 5.0        # step 2: no adjustment at a ratio of 5 or below
LHN_RATIO_SLOPE = 0.95       # ... then 0.95% of baseline per 1% above 5


def _ods_rows(path, sheet_name):
    """Minimal ODS reader — rows of strings for one sheet.

    Written against the stdlib rather than adding odfpy: an ODS is a zip of
    XML, and one sheet of a published statistics table needs none of the
    library's features. Repeated cells are expanded (the format runs them),
    capped so a trailing 'repeat 16384 empty columns' cannot blow up a row.
    """
    import zipfile
    import xml.etree.ElementTree as ET
    T = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    TX = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    OF = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    with zipfile.ZipFile(path) as z:
        root = ET.fromstring(z.read("content.xml"))
    for t in root.iter("{%s}table" % T):
        if t.get("{%s}name" % T) != sheet_name:
            continue
        out = []
        for row in t.findall("{%s}table-row" % T):
            cells = []
            for c in row.findall("{%s}table-cell" % T):
                rep = int(c.get("{%s}number-columns-repeated" % T, 1))
                v = c.get("{%s}value" % OF)
                if v is None:
                    v = "".join("".join(p.itertext())
                                for p in c.findall("{%s}p" % TX))
                cells += [v] * min(rep, 64)
            out.append(cells)
        return out
    return []


# Authorities created by local government reorganisation, with the date they
# came into being. The planning platform re-points a predecessor district's
# plan at its successor, so without this a predecessor's requirement reads as
# the whole new authority's: Somerset was showing the Sedgemoor Local Plan's
# 644 homes a year against a county need of 3,784 — a fabricated 83%
# shortfall, and North Northamptonshire the same from a Joint Core Strategy
# record holding 7,000 homes where the plan's own figure is around 35,000.
# A plan whose period begins before its authority existed cannot speak for
# that authority, which catches these without judging any of them by name.
REORGANISED_AUTHORITIES = {
    "E06000058": 2019, "E06000059": 2019, "E07000244": 2019,   # BCP, Dorset, E Suffolk
    "E07000245": 2019, "E07000246": 2019,                      # W Suffolk, Somerset W & Taunton
    "E06000060": 2020,                                          # Buckinghamshire
    "E06000061": 2021, "E06000062": 2021,                      # N & W Northamptonshire
    "E06000063": 2023, "E06000064": 2023,                      # Cumberland, Westmorland & Furness
    "E06000065": 2023, "E06000066": 2023,                      # North Yorkshire, Somerset
}


def build_housing_need():
    """Standard-method local housing need per authority — dataset housing_need.

    NPPF (August 2026) Annex D specifies the standard method completely, and
    both of its inputs are published free:

      step 1  baseline = 0.8% of the authority's existing dwelling stock
              (MHCLG Live Table 125, dwelling stock by local authority)
      step 2  adjusted for affordability. No adjustment where the ratio is 5
              or below; for each 1% above 5, the baseline rises 0.95%:
                  factor = ((5yr average ratio - 5) / 5) x 0.95 + 1
              (ONS median WORKPLACE-BASED house-price-to-earnings ratio by
              local authority — the series the Framework names, not our
              neighbourhood affordability layer. ONS publishes the five-year
              average as its own column, which is exactly the mean Annex D
              asks for.)

    That yields the minimum annual housing need figure that governs how much
    land an authority has to find — the number the rest of the housing policy
    machinery hangs off.

    The builder then sets each authority's ADOPTED plan requirement against
    it, via the same three planning-platform CSVs the local_plan_housing
    dataset uses, joined plan -> organisation -> local authority district.
    Annex D para 9(c) attaches a 20% buffer where an annual average
    requirement is "80% or less of the most up-to-date local housing need
    figure" — but only for a plan examined against a pre-December-2024
    Framework, a limb we cannot confirm per authority. So the flag is named
    below80 and described as one limb of that test, never as the buffer
    itself."""
    key = "housing_need"
    bpath, bhow = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries to join housing need onto")
        _note(key, "no LAD boundaries")
        return {}

    def _num(v):
        try:
            f = float(str(v).replace(",", "").strip())
            return f if f == f else None
        except (TypeError, ValueError):
            return None

    # ---- step 1: dwelling stock -------------------------------------------
    spath, show = _resolve_source(key, ["DWELLING_STOCK_SRC"],
                                  DWELLING_STOCK_URL, "dwelling-stock.ods")
    stock, stock_year = {}, None
    if spath is not None and str(spath).lower().endswith(".ods"):
        rows = _ods_rows(spath, "LT_125_unrounded") or _ods_rows(spath, "LT_125_rounded")
        hdr_i = next((i for i, r in enumerate(rows[:12])
                      if any("new ons code" in str(c).lower() for c in r)), None)
        if hdr_i is None:
            _warn(key, "could not find the header row in Live Table 125")
        else:
            hdr = rows[hdr_i]
            # Rightmost column whose header names a year and which actually
            # carries figures — the table keeps trailing blank columns, and
            # the newest year is provisional ("2025 [p]").
            years = [(i, re.match(r"\s*(\d{4})", str(c)).group(1))
                     for i, c in enumerate(hdr)
                     if re.match(r"\s*\d{4}", str(c))]
            for i, yr in reversed(years):
                vals = {}
                for r in rows[hdr_i + 1:]:
                    if len(r) <= max(i, 1):
                        continue
                    code = str(r[1]).strip()
                    v = _num(r[i])
                    if re.fullmatch(r"E0[6-9]\d{6}", code) and v:
                        vals[code] = v
                if len(vals) >= 200:
                    stock, stock_year = vals, yr
                    break
            if not stock:
                _warn(key, "no populated year column in Live Table 125")
    elif spath is not None:
        # Drop-in CSV: any LAD-code column plus a stock column.
        with open(spath, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                code = next((str(v).strip() for v in r.values()
                             if re.fullmatch(r"E0[6-9]\d{6}", str(v).strip())), None)
                v = next((_num(r[k]) for k in r
                          if "stock" in k.lower() or "dwelling" in k.lower()), None)
                if code and v:
                    stock[code] = v
    if not stock:
        _warn(key, f"no dwelling stock figures ({show}) — set DWELLING_STOCK_SRC "
                   "to MHCLG Live Table 125 (ODS) or a LAD-code CSV")
        _note(key, "no dwelling stock")
        return {}

    # ---- step 2: affordability ratio --------------------------------------
    apath, ahow = _resolve_source(key, ["AFFORDABILITY_RATIO_SRC"],
                                  AFFORDABILITY_URL, "affordability-ratio.xlsx")
    ratio = {}
    if apath is not None and str(apath).lower().endswith((".xlsx", ".xlsm")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(apath, read_only=True, data_only=True)
            # Table 5c: median affordability ratio by local authority district.
            ws = wb["5c"] if "5c" in wb.sheetnames else wb[wb.sheetnames[-1]]
            hdr, avg_i, code_i = None, None, None
            for r in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c).strip() for c in r]
                if hdr is None:
                    if any(c.lower().startswith("local authority code") for c in cells):
                        hdr = cells
                        code_i = next(i for i, c in enumerate(cells)
                                      if c.lower().startswith("local authority code"))
                        # ONS publishes the five-year average as its own
                        # column — the exact mean Annex D asks for.
                        avg_i = next((i for i, c in enumerate(cells)
                                      if "5-year average" in c.lower()
                                      or "5 year average" in c.lower()), None)
                    continue
                if code_i is None or avg_i is None or len(cells) <= max(code_i, avg_i):
                    continue
                code, v = cells[code_i], _num(cells[avg_i])
                if re.fullmatch(r"E0[6-9]\d{6}", code) and v:
                    ratio[code] = v
        except Exception as exc:                       # noqa: BLE001
            _warn(key, f"could not read the affordability workbook: {exc}")
    elif apath is not None:
        with open(apath, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                code = next((str(v).strip() for v in r.values()
                             if re.fullmatch(r"E0[6-9]\d{6}", str(v).strip())), None)
                v = next((_num(r[k]) for k in r if "ratio" in k.lower()), None)
                if code and v:
                    ratio[code] = v
    if not ratio:
        # Without step 2 the baseline still stands, but the figure would be a
        # different quantity from the one the Framework defines. Say so rather
        # than publishing an unadjusted number as "local housing need".
        _warn(key, f"no affordability ratios ({ahow}) — housing need cannot be "
                   "computed without step 2 of the standard method")
        _note(key, "no affordability ratios")
        return {}

    # ---- the adopted plan requirement, per authority ----------------------
    # lad -> every plan carrying a requirement for it, and plan -> the
    # authorities it covers. BOTH directions are needed, because the platform
    # models two situations that look identical from one side and mean
    # opposite things (see the guard below).
    plan_by_lad, plan_areas = {}, {}
    ppath, _phow = _resolve_source(key, ["LOCAL_PLAN_SRC"],
                                   PLANNING_DATA_BASE + "local-plan.csv",
                                   "local-plan.csv")
    hpath, _hhow = _resolve_source(key, ["LOCAL_PLAN_HOUSING_SRC"],
                                   PLANNING_DATA_BASE + "local-plan-housing.csv",
                                   "local-plan-housing.csv")
    opath, _ohow = _resolve_source(key, ["LOCAL_AUTHORITY_SRC"],
                                   PLANNING_DATA_BASE + "local-authority.csv",
                                   "local-authority.csv")
    if ppath and hpath and opath:
        orgs = {}
        with open(opath, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                orgs[r.get("entity", "")] = (
                    r.get("local-authority-district") or "").strip()
        hous = {}
        with open(hpath, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                hous[r.get("local-plan", "")] = r
        with open(ppath, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                lad = orgs.get(r.get("organisation-entity", ""), "")
                h = hous.get(r.get("reference", ""))
                if not lad or not h:
                    continue
                req = _num(h.get("required-housing"))
                ps = (r.get("period-start-date") or "")[:4]
                pe = (r.get("period-end-date") or "")[:4]
                if not req or not (ps.isdigit() and pe.isdigit()):
                    continue
                yrs = int(pe) - int(ps)
                if yrs <= 0:
                    continue
                cand = {"required": int(req), "years": yrs,
                        "annual": req / yrs, "start": ps, "end": pe,
                        "ref": r.get("reference", ""),
                        "name": (r.get("name") or "").strip() or None,
                        "adopted": (r.get("adopted-date") or "")[:10] or None}
                plan_by_lad.setdefault(lad, []).append(cand)
                plan_areas.setdefault(cand["ref"], set()).add(lad)

    n_lhn = n_plan = n_below = 0

    def hn_props(f):
        nonlocal n_lhn, n_plan, n_below
        code = str(_cell(f, code_col) or "").strip()
        st, ra = stock.get(code), ratio.get(code)
        if st is None or ra is None:
            return {"lad_code": code, "status": "no_data",
                    "src": "outside the England-only sources for the standard "
                           "method (dwelling stock and affordability ratio)"}
        baseline = st * LHN_STOCK_PCT
        factor = (1.0 if ra <= LHN_RATIO_FLOOR
                  else ((ra - LHN_RATIO_FLOOR) / LHN_RATIO_FLOOR)
                       * LHN_RATIO_SLOPE + 1.0)
        lhn = baseline * factor
        n_lhn += 1
        out = {
            "lad_code": code, "status": "ok",
            "stock": int(round(st)), "stock_year": stock_year,
            "afford_ratio": round(ra, 2),
            "afford_factor": round(factor, 3),
            "baseline": int(round(baseline)),
            "lhn": int(round(lhn)),
            "src": f"NPPF (Aug 2026) Annex D standard method: "
                   f"{LHN_STOCK_PCT * 100:g}% of {int(round(st)):,} dwellings "
                   f"({stock_year}) x {factor:.3f} affordability adjustment "
                   f"(5yr median workplace-based ratio {ra:g})",
        }
        # ---- the plan comparison, and the two ways it goes wrong ----------
        # A ratio is only stated where ONE plan covers this authority and that
        # plan covers ONLY this authority. The two exclusions are not fussiness:
        #
        #  * Several plans -> local government reorganisation. Dorset (2019)
        #    carries three predecessor district plans and North Yorkshire
        #    (2023) carries seven. Picking one and calling it the county's
        #    requirement produced 186 and 315 homes a year against needs of
        #    3,273 and 4,173 — a fabricated 95% shortfall. Summing them instead
        #    would assume the plans partition the area rather than supersede
        #    one another, which the data does not say.
        #  * One plan across several authorities -> a joint plan. The Greater
        #    Norwich Local Plan's 40,541 homes belong to three authorities
        #    together; charging the whole figure to each put Norwich at 262% of
        #    its need. There is no published split to apportion by.
        #
        # Both cases still carry the plan count, because "no single plan for
        # this authority" is itself a strong signal — an authority running on
        # predecessor plans almost certainly has no up-to-date one.
        plans = plan_by_lad.get(code) or []
        if plans and lhn > 0:
            out["plan_count"] = len(plans)
            shared = sorted({a for p in plans
                             for a in (plan_areas.get(p["ref"]) or set())
                             if a != code})
            if len(plans) > 1:
                out["plan_status"] = "multiple_plans"
                out["plan_note"] = (
                    f"{len(plans)} separate plans carry a requirement for this "
                    "authority — predecessor districts of a reorganised area. "
                    "No single figure represents it, so no comparison is shown.")
            elif shared:
                out["plan_status"] = "joint_plan"
                out["plan_name"] = plans[0]["name"]
                out["plan_period"] = f"{plans[0]['start']}-{plans[0]['end']}"
                out["plan_note"] = (
                    f"a joint plan shared with {len(shared)} other "
                    f"{'authority' if len(shared) == 1 else 'authorities'} — "
                    "its requirement covers them together and there is no "
                    "published split, so no comparison is shown.")
            elif (code in REORGANISED_AUTHORITIES
                  and plans[0]["start"].isdigit()
                  and int(plans[0]["start"]) < REORGANISED_AUTHORITIES[code]):
                out["plan_status"] = "predecessor_plan"
                out["plan_name"] = plans[0]["name"]
                out["plan_period"] = f"{plans[0]['start']}-{plans[0]['end']}"
                out["plan_note"] = (
                    f"the only plan on record began in {plans[0]['start']}, "
                    f"before this authority was created in "
                    f"{REORGANISED_AUTHORITIES[code]} — it speaks for a "
                    "predecessor district, not the whole area, so no "
                    "comparison is shown.")
            else:
                p = plans[0]
                n_plan += 1
                pct = 100.0 * p["annual"] / lhn
                out.update({
                    "plan_status": "compared",
                    "plan_required": p["required"], "plan_years": p["years"],
                    "plan_annual": int(round(p["annual"])),
                    "plan_period": f"{p['start']}-{p['end']}",
                    "plan_name": p["name"], "plan_adopted": p["adopted"],
                    "plan_vs_lhn": round(pct, 1),
                })
                if pct <= 80:
                    n_below += 1
                    # Named for what it is — ONE limb of the Annex D 9(c)
                    # test. The other limb (a requirement adopted in the last
                    # five years, examined against a pre-December-2024
                    # Framework) cannot be confirmed per authority from this
                    # data, so this is not the 20% buffer itself.
                    out["below80"] = True
        return out

    gdf = gpd.read_file(bpath)
    code_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*cd", str(c).lower()):
            code_col = c
            break
    code_col = code_col or _find_col(gdf, ["lad_code", "code", "areacd"],
                                     contains=True)
    name_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*nm", str(c).lower()):
            name_col = c
            break
    if code_col is None:
        _warn(key, "no LAD code column on the boundary file")
        _note(key, "no LAD code column")
        return {}

    rows = _emit(gdf, key, name_col=name_col, id_col=code_col, want="polygon",
                 props_fn=hn_props)
    print(f"  [{key}] stock {len(stock)} authorities ({stock_year}), "
          f"ratios {len(ratio)}; {n_lhn} with a housing need figure, "
          f"{n_plan} comparable against a single adopted plan requirement, "
          f"{n_below} of those sit at or below 80% of need")
    _note(key, f"{show}; {n_lhn} authorities with standard-method need", len(rows))
    return {key: rows}


# ONS Output Area 2021 boundaries, generalised & clipped. 188,880 polygons,
# paged 2,000 at a time — the finest census geography the connectivity metric
# publishes.
OA_BOUNDARY_URL = ("https://services1.arcgis.com/ESMARspQHYMw9BZ9/arcgis/rest/"
                   "services/Output_Areas_2021_EW_BGC_V2/FeatureServer/0/query"
                   "?where=1%3D1&outFields=OA21CD,LSOA21CD&outSR=4326&f=geojson")

# DfT transport connectivity metric (2025). One 68 MB ODS holding 1.09 GB of
# XML: Metadata, then OA (188,884 rows), LSOA (35,672), LAD (331), RGN (10).
CONNECTIVITY_URL = ("https://assets.publishing.service.gov.uk/media/"
                    "68c966fc07d9e92bc5517b80/connectivity_metrics_2025.ods")

# Column name -> compact prop key. Modes: w walking, c cycling, p public
# transport, d driving, a all modes. Purposes: emp edu hea lei sho res, all.
# Storing all 35 keeps the mode/purpose switch on the client free — it is a
# repaint of props already loaded, never a refetch (the price-choropleth
# lesson). Short keys because 35 of them ride on every one of 33,755 rows.
CONNECTIVITY_MODES = {
    "walking": "w", "cycling": "c", "public transport": "p",
    "driving": "d", "overall": "a",
}
CONNECTIVITY_PURPOSES = {
    "employment": "emp",
    # The public-transport block calls this column "Business"; same measure.
    "business": "emp",
    "education": "edu", "healthcare": "hea",
    "leisure and community": "lei", "shopping": "sho",
    "residential": "res", "overall": "all",
}


def _connectivity_key(header):
    """'Healthcare (public transport)' -> 'p_hea'; bare 'Overall' -> 'a_all'."""
    h = str(header).strip()
    m = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", h)
    if m:
        purpose, mode = m.group(1).strip().lower(), m.group(2).strip().lower()
    else:
        # The all-modes block drops the suffix: "Overall", "Employment (overall)"
        purpose, mode = h.lower(), "overall"
    p, mo = CONNECTIVITY_PURPOSES.get(purpose), CONNECTIVITY_MODES.get(mode)
    return f"{mo}_{p}" if p and mo else None


def _ods_sheet_rows(path, wanted):
    """Stream one or more sheets out of a large ODS.

    iterparse rather than a whole-document parse: content.xml here is 1.09 GB
    and ElementTree runs out of memory on it. Rows are cleared AND detached
    from the root as they close, which is what keeps memory flat — clearing
    the row alone still leaves it hanging off the tree. Yields (sheet, cells).
    """
    import xml.etree.ElementTree as ET
    T = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    TX = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    OF = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    import zipfile
    z = zipfile.ZipFile(path)
    sheet = None
    with z.open("content.xml") as fh:
        ctx = ET.iterparse(fh, events=("start", "end"))
        _, root = next(ctx)
        for ev, el in ctx:
            if ev == "start" and el.tag == "{%s}table" % T:
                sheet = el.get("{%s}name" % T)
            elif ev == "end" and el.tag == "{%s}table-row" % T:
                if sheet in wanted:
                    cells = []
                    for c in el.findall("{%s}table-cell" % T):
                        rep = int(c.get("{%s}number-columns-repeated" % T, 1))
                        v = c.get("{%s}value" % OF)
                        if v is None:
                            v = "".join("".join(p.itertext())
                                        for p in c.findall("{%s}p" % TX))
                        cells += [v] * min(rep, 64)
                    while cells and cells[-1] == "":
                        cells.pop()
                    if cells:
                        yield sheet, cells
                el.clear()
                root.clear()
            elif ev == "end" and el.tag == "{%s}table" % T:
                el.clear()
                root.clear()


def build_connectivity():
    """DfT connectivity scores on LSOA and LAD polygons — datasets
    connectivity_lsoa and connectivity_lad.

    How easily residents can reach employment, education, healthcare, leisure,
    shopping and other homes, by walking, cycling, public transport and
    driving. Scores are 0-100, higher is better connected, computed by DfT from
    real networks and timetables.

    Two reasons this is worth more than another choropleth. It is a NATIONAL
    equivalent of PTAL, which we can only show for London because TfL's is
    London-only. And NPPF (Aug 2026) policies TR3(2) and S5(3) name the
    Connectivity Tool by URL as the thing that "should be used ... in assessing
    the connectivity of particular locations proposed for development" — so
    this is the Framework's own measure, not a proxy we chose.

    All 35 scores ride on every row so the client's mode/purpose switch is a
    repaint rather than a refetch. The LAD band exists because 35,672 LSOAs are
    sub-pixel at national zoom and features_in_bbox drops sub-pixel features;
    it is derived here by POPULATION-WEIGHTING the LSOA scores rather than read
    from the file's own LAD sheet, which is keyed by name on 2022 geography and
    so predates Cumberland, North Yorkshire and Somerset. Aggregating ourselves
    avoids both the name join and the vintage mismatch.

    NOTE the file is badged Experimental Statistics / prepublication, covers
    Q4 2024, and states its own licence as 'TBA' against the GOV.UK page's
    default OGL v3. Both facts are carried into the layer copy."""
    key = "connectivity_lsoa"
    fpath, fhow = _resolve_source(key, ["CONNECTIVITY_SRC"], CONNECTIVITY_URL,
                                  "connectivity_metrics.ods")
    if fpath is None:
        _warn(key, "no connectivity ODS — set CONNECTIVITY_SRC")
        _note(key, "no source file")
        return {}

    lpath, lhow = _resolve_source("lsoa_boundary", ["LSOA_BOUNDARY_SRC"], None,
                                  "lsoa-boundaries.geojson")
    if lpath is None:
        committed = ROOT / "web" / "data" / "lsoa_imd.geojson"
        if committed.exists():
            lpath, lhow = committed, "committed app layer"
        else:
            _warn(key, "no LSOA polygons to join connectivity onto")
            _note(key, "no LSOA boundaries")
            return {}

    print(f"  [{key}] streaming the connectivity workbook "
          f"(1 GB of XML — a couple of minutes) ...")
    cols, scores = None, {}
    for sheet, cells in _ods_sheet_rows(fpath, {"LSOA"}):
        if cols is None:
            if str(cells[0]).strip().upper().startswith("LSOA"):
                cols = [_connectivity_key(h) for h in cells[1:]]
            continue
        code = str(cells[0]).strip()
        if not re.fullmatch(r"[EW]01\d{6}", code):
            continue
        row = {}
        for k, v in zip(cols, cells[1:]):
            if not k:
                continue
            try:
                row[k] = round(float(v), 1)
            except (TypeError, ValueError):
                pass
        if row:
            scores[code] = row
    if not scores:
        _warn(key, "no LSOA rows parsed from the connectivity workbook")
        _note(key, "no rows parsed")
        return {}
    print(f"  [{key}] {len(scores):,} LSOAs scored, "
          f"{len(scores[next(iter(scores))])} metrics each")

    gdf = gpd.read_file(lpath)
    code_col = _find_col(gdf, ["lsoa_code", "lsoa21cd", "lsoa11cd", "code"],
                         contains=True)
    name_col = _find_col(gdf, ["lsoa_name", "name"], contains=True)
    pop_col = _find_col(gdf, ["population", "pop"], contains=True)
    if code_col is None:
        _warn(key, f"no LSOA code column, got {list(gdf.columns)[:12]}")
        _note(key, "no code column")
        return {}

    n_hit, n_miss = 0, 0

    def conn_props(f):
        nonlocal n_hit, n_miss
        code = str(_cell(f, code_col) or "").strip()
        row = scores.get(code)
        if row is None:
            n_miss += 1
            return {"lsoa": code, "status": "no_score"}
        n_hit += 1
        return dict({"lsoa": code}, **row)

    rows = _emit(gdf, key, name_col=name_col, id_col=code_col, want="polygon",
                 props_fn=conn_props)
    print(f"  [{key}] {n_hit:,} polygons matched a score, {n_miss:,} did not")

    # ---- the LAD band, population-weighted from the LSOA scores -----------
    out = {key: rows}
    bpath, _bhow = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries — the wide-zoom band will be missing")
        _note(key, f"{fhow}; {n_hit} LSOAs scored", len(rows))
        return out

    lads = gpd.read_file(bpath)
    lad_code_col = None
    for c in lads.columns:
        if re.fullmatch(r"lad\d*cd", str(c).lower()):
            lad_code_col = c
            break
    lad_code_col = lad_code_col or _find_col(lads, ["lad_code", "code"],
                                             contains=True)
    lad_name_col = None
    for c in lads.columns:
        if re.fullmatch(r"lad\d*nm", str(c).lower()):
            lad_name_col = c
            break
    if lad_code_col is None:
        _warn("connectivity_lad", "no LAD code column")
        _note(key, f"{fhow}; {n_hit} LSOAs scored", len(rows))
        return out

    # Centroid-in-polygon rather than a name join: LSOA and LAD names come
    # from different vintages, and matching them by hand is exactly the
    # failure mode this dataset is trying to avoid.
    pts = gdf[[code_col] + ([pop_col] if pop_col else [])].copy()
    pts["geometry"] = gdf.geometry.representative_point()
    pts = gpd.GeoDataFrame(pts, geometry="geometry", crs=gdf.crs)
    if lads.crs is not None and pts.crs is not None and lads.crs != pts.crs:
        pts = pts.to_crs(lads.crs)
    joined = gpd.sjoin(pts, lads[[lad_code_col, "geometry"]],
                       how="inner", predicate="within")

    agg = {}
    for _, r in joined.iterrows():
        row = scores.get(str(r[code_col]).strip())
        if not row:
            continue
        lad = str(r[lad_code_col]).strip()
        w = 1.0
        if pop_col:
            try:
                w = max(0.0, float(r[pop_col]))
            except (TypeError, ValueError):
                w = 0.0
        if w <= 0:
            w = 1.0
        acc = agg.setdefault(lad, {"w": 0.0, "n": 0, "v": {}})
        acc["w"] += w
        acc["n"] += 1
        for k, v in row.items():
            acc["v"][k] = acc["v"].get(k, 0.0) + v * w

    def lad_props(f):
        code = str(_cell(f, lad_code_col) or "").strip()
        a = agg.get(code)
        if not a or a["w"] <= 0:
            return {"lad_code": code, "status": "no_score"}
        p = {k: round(v / a["w"], 1) for k, v in a["v"].items()}
        p.update({"lad_code": code, "n_lsoa": a["n"],
                  "weighted": bool(pop_col)})
        return p

    lad_rows = _emit(lads, "connectivity_lad", name_col=lad_name_col,
                     id_col=lad_code_col, want="polygon", props_fn=lad_props)
    print(f"  [connectivity_lad] {len(agg):,} authorities aggregated from "
          f"{len(joined):,} LSOAs"
          f"{' (population-weighted)' if pop_col else ' (unweighted mean)'}")
    out["connectivity_lad"] = lad_rows
    _note(key, f"{fhow}; {n_hit} LSOAs scored, {len(agg)} authorities",
          len(rows) + len(lad_rows))
    return out


def build_connectivity_oa():
    """DfT connectivity at OUTPUT AREA level — dataset connectivity_oa.

    188,884 Output Areas against 35,672 LSOAs: roughly 5.6x the detail, an OA
    being about 125 households. This is as fine as the published data goes —
    the workbook's own metadata states "lowest geography: Output Area (OA
    2021)", so the 100 m grid DfT's interactive tool draws is rendered inside
    that tool and is not distributed.

    This dataset exists to be TILED, not queried per viewport. The bbox RPC
    spends roughly 400 microseconds per feature assembling JSON, which is what
    made PTAL's 159,451 cells unservable at a London-wide zoom (52,301 cells,
    14 MB, 20.8 s), and 188,884 polygons would hit the same wall. It is loaded
    into map_features so build_ptal_tiles.py can export it to PMTiles — and,
    as with PTAL, the table copy keeps earning its place because the deep dive
    and spot summary query it point-by-point, which is cheap and lets a site
    report quote the connectivity of the actual site rather than its
    neighbourhood average."""
    key = "connectivity_oa"
    fpath, fhow = _resolve_source(key, ["CONNECTIVITY_SRC"], CONNECTIVITY_URL,
                                  "connectivity_metrics.ods")
    if fpath is None:
        _warn(key, "no connectivity ODS — set CONNECTIVITY_SRC")
        _note(key, "no source file")
        return {}
    bpath, bhow = _resolve_arcgis_source(key, ["OA_BOUNDARY_SRC"],
                                         OA_BOUNDARY_URL, "oa-boundaries.geojson")
    if bpath is None:
        _warn(key, f"no Output Area boundaries ({bhow})")
        _note(key, "no OA boundaries")
        return {}

    print(f"  [{key}] streaming the OA sheet (188,884 rows) ...")
    cols, scores = None, {}
    for _sheet, cells in _ods_sheet_rows(fpath, {"OA"}):
        if cols is None:
            if str(cells[0]).strip().upper().startswith("OA"):
                cols = [_connectivity_key(h) for h in cells[1:]]
            continue
        code = str(cells[0]).strip()
        if not re.fullmatch(r"[EW]00\d{6}", code):
            continue
        row = {}
        for k, v in zip(cols, cells[1:]):
            if not k:
                continue
            try:
                row[k] = round(float(v), 1)
            except (TypeError, ValueError):
                pass
        if row:
            scores[code] = row
    if not scores:
        _warn(key, "no OA rows parsed from the connectivity workbook")
        _note(key, "no rows parsed")
        return {}
    print(f"  [{key}] {len(scores):,} Output Areas scored")

    gdf = gpd.read_file(bpath)
    code_col = _find_col(gdf, ["oa21cd", "oa_code", "code"], contains=True)
    if code_col is None:
        _warn(key, f"no OA code column, got {list(gdf.columns)[:12]}")
        _note(key, "no code column")
        return {}
    lsoa_col = _find_col(gdf, ["lsoa21cd"], contains=True)

    n_hit, n_miss = 0, 0

    def oa_props(f):
        nonlocal n_hit, n_miss
        code = str(_cell(f, code_col) or "").strip()
        row = scores.get(code)
        if row is None:
            n_miss += 1
            return {"oa": code, "status": "no_score"}
        n_hit += 1
        out = {"oa": code}
        if lsoa_col is not None:
            out["lsoa"] = str(_cell(f, lsoa_col) or "").strip() or None
        return dict(out, **row)

    rows = _emit(gdf, key, name_col=code_col, id_col=code_col, want="polygon",
                 props_fn=oa_props)
    print(f"  [{key}] {n_hit:,} polygons matched a score, {n_miss:,} did not")
    _note(key, f"{fhow}; {n_hit} Output Areas scored", len(rows))
    return {key: rows}


def build_income():
    """Household earnings on LAD polygons — dataset lad_income.

    Source: ONS ASHE Table 8 (residence-based gross annual pay, full-time, by
    local authority). No stable download URL, so the usual pattern: ASHE_SRC
    (URL or path) or the drop-in data/raw/ashe-la-earnings.csv. Expected
    columns: an LAD code column and a median annual pay column; extra columns
    are ignored. The affordability ratio itself (median price / income) is
    computed in-DB by rebuild_lad_affordability() after each PPD load, so this
    builder only ships the income side."""
    key = "lad_income"
    # Default: the NOMIS API (stable, parameterised — not a hashed per-release
    # link) serving ASHE resident analysis, median gross ANNUAL pay, full-time,
    # by local authority. The dimension codes are best-effort: if they drift,
    # the run warns and skips — and a wrong pay dimension cannot load silently,
    # because the >£1,000 sanity guard below drops hourly (~£15) and weekly
    # (~£700) figures wholesale. ASHE_SRC overrides with a hand-picked CSV.
    ashe_default = ("https://www.nomisweb.co.uk/api/v01/dataset/NM_30_1.data.csv"
                    "?geography=TYPE434&date=latest&sex=8&item=2&pay=7"
                    "&measures=20100"
                    "&select=geography_code,geography_name,date_name,obs_value")
    ipath, ihow = _resolve_source(key, ["ASHE_SRC"], ashe_default,
                                  "ashe-la-earnings.csv")
    if ipath is None:
        _warn(key, "no earnings CSV — download ONS ASHE Table 8 (residence, "
                   "gross annual pay, full-time) and supply it as ASHE_SRC or "
                   "drop in data/raw/ashe-la-earnings.csv")
        _note(key, "ASHE_SRC not set and NOMIS default failed")
        return {}
    bpath, bhow = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries to join earnings onto")
        _note(key, "no LAD boundaries")
        return {}
    print(f"  [{key}] reading {ipath.name} ({ihow}) ...")
    df = _read_csv(ipath)
    ccol = _find_col(df, ["area code", "areacd", "ons code", "lad code",
                          "ladcd", "geography code", "code"], contains=True)
    mcol = _find_col(df, ["median"], contains=True) \
        or _find_col(df, ["obs_value", "annual pay", "gross annual", "pay"],
                     contains=True)
    ycol = _find_col(df, ["date_name", "year", "period", "date", "asof"],
                     contains=True)
    if ccol is None or mcol is None:
        _warn(key, f"need LAD-code + median-pay columns, got "
                   f"{list(df.columns)[:12]}")
        _note(key, "unrecognised earnings CSV")
        return {}
    income = {}
    asof = ""
    for _, row in df.iterrows():
        code = str(row[ccol]).strip()
        v = _num(row, mcol)
        # ASHE publishes suppressed cells as 'x'/':' — _num returns None; skip.
        if code and v and v > 1000:
            income[code] = round(v)
            if ycol is not None and not asof:
                asof = str(row[ycol]).strip()

    gdf = gpd.read_file(bpath)
    code_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*cd", str(c).lower()):
            code_col = c
            break
    code_col = code_col or _find_col(gdf, ["lad_code", "code", "areacd"],
                                     contains=True)
    name_col = None
    for c in gdf.columns:
        if re.fullmatch(r"lad\d*nm", str(c).lower()):
            name_col = c
            break
    keep = gdf[gdf[code_col].astype(str).str.strip().isin(income)].copy()
    print(f"  [{key}] {len(keep)}/{len(gdf)} authorities matched an income")

    rows = _emit(keep, key, name_col=name_col, id_col=code_col, want="polygon",
                 props_fn=lambda f: {
                     "lad_code": str(_cell(f, code_col)).strip(),
                     "income_median": income.get(str(_cell(f, code_col)).strip()),
                     "asof": asof or None})
    _note(key, ihow, len(rows))
    return {key: rows}


def build_msoa_income():
    """Neighbourhood-scale household income — dataset msoa_income.

    MSOA is the finest geography with an OFFICIAL income estimate: ONS small
    area income estimates, ~7,200 areas of roughly 4,000 households each,
    against ~360 local authorities. This is what stops a market town's
    affordability being written by the city an hour away — the LAD layer keeps
    the wide-zoom view, this one takes over close in.

    Uses HOUSEHOLD income (the small-area series is household, not individual
    pay) — also the better affordability denominator. The ONS file has no
    stable URL, so: MSOA_INCOME_SRC (URL or path) or the drop-in
    data/raw/msoa-income.csv, expected to carry an MSOA code column and a
    total/net annual household income column. Values that look WEEKLY
    (median < £2,000) are converted to annual x52, because ONS publishes both
    shapes and the difference is a 52x error if trusted blindly."""
    key = "msoa_income"
    ipath, ihow = _resolve_source(key, ["MSOA_INCOME_SRC"], None,
                                  "msoa-income.csv")
    if ipath is None:
        _warn(key, "no MSOA income CSV — download the ONS 'small area income "
                   "estimates' MSOA table (search ons.gov.uk) and supply it as "
                   "MSOA_INCOME_SRC or drop in data/raw/msoa-income.csv. The "
                   "LAD-level affordability layer still works without it.")
        _note(key, "MSOA_INCOME_SRC not set")
        return {}
    bpath, bhow = _resolve_arcgis_source(
        key, ["MSOA_BOUNDARIES_SRC"], MSOA_DEFAULT_URL, "msoa-boundaries.geojson")
    if bpath is None:
        _warn(key, "no MSOA boundaries to join income onto — set "
                   "MSOA_BOUNDARIES_SRC to an ONS Open Geography MSOA 2021 "
                   "BGC GeoJSON/FeatureServer URL")
        _note(key, "no MSOA boundaries")
        return {}
    print(f"  [{key}] reading {ipath.name} ({ihow}) ...")
    df = _read_csv(ipath)
    ccol = _find_col(df, ["msoa code", "msoa21cd", "msoa11cd", "geography code",
                          "area code", "code"], contains=True)
    mcol = _find_col(df, ["total annual income", "net annual income",
                          "annual income", "total weekly income",
                          "net weekly income", "weekly income", "income"],
                     contains=True)
    ycol = _find_col(df, ["date_name", "year", "period", "date"], contains=True)
    if ccol is None or mcol is None:
        _warn(key, f"need MSOA-code + income columns, got "
                   f"{list(df.columns)[:12]}")
        _note(key, "unrecognised income CSV")
        return {}
    vals = []
    raw = {}
    asof = ""
    for _, row in df.iterrows():
        code = str(row[ccol]).strip()
        v = _num(row, mcol)
        if code and v and v > 0:
            raw[code] = v
            vals.append(v)
            if ycol is not None and not asof:
                asof = str(row[ycol]).strip()
    if not raw:
        _warn(key, "no usable income rows")
        _note(key, "no usable rows")
        return {}
    vals.sort()
    weekly = vals[len(vals) // 2] < 2000
    if weekly:
        print(f"  [{key}] values look WEEKLY (median £{vals[len(vals)//2]:.0f})"
              f" — converting to annual x52")
    income = {c: round(v * 52 if weekly else v) for c, v in raw.items()}

    gdf = gpd.read_file(bpath)
    code_col = None
    for c in gdf.columns:
        if re.fullmatch(r"msoa\d*cd", str(c).lower()):
            code_col = c
            break
    code_col = code_col or _find_col(gdf, ["msoa_code", "code", "areacd"],
                                     contains=True)
    name_col = None
    for c in gdf.columns:
        if re.fullmatch(r"msoa\d*nm", str(c).lower()):
            name_col = c
            break
    keep = gdf[gdf[code_col].astype(str).str.strip().isin(income)].copy()
    print(f"  [{key}] {len(keep)}/{len(gdf)} MSOAs matched an income")

    rows = _emit(keep, key, name_col=name_col, id_col=code_col, want="polygon",
                 props_fn=lambda f: {
                     "msoa_code": str(_cell(f, code_col)).strip(),
                     "income_median": income.get(str(_cell(f, code_col)).strip()),
                     "hh": True,           # household income, not individual pay
                     "asof": asof or None})
    _note(key, f"{ihow}{' (weekly x52)' if weekly else ''}", len(rows))
    return {key: rows}


CENSUS_TS062_URL = "https://www.nomisweb.co.uk/output/census/2021/census2021-ts062.zip"


def build_census_students():
    """Census 2021 full-time students (NS-SeC class L15) per local authority
    — the PBSA demand base. NOMIS bulk zip, no key; the ltla CSV inside is
    joined onto LAD polygons."""
    key = "census_students"
    path, how = _resolve_source(key, ["CENSUS_STUDENTS_SRC"],
                                CENSUS_TS062_URL, "census-ts062.zip")
    if path is None:
        _warn(key, f"{how} — set CENSUS_STUDENTS_SRC to the NOMIS "
                   "census2021-ts062 bulk zip")
        _note(key, how)
        return {}
    bpath, _b = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries to join students onto")
        _note(key, "no LAD boundaries")
        return {}
    import zipfile as _zf
    import io as _io
    by_code = {}
    with _zf.ZipFile(path) as z:
        member = next((m for m in z.namelist()
                       if "ltla" in m.lower() and m.lower().endswith(".csv")),
                      None)
        if member is None:
            _warn(key, f"no ltla CSV inside {path.name} "
                       f"(members: {z.namelist()[:6]})")
            _note(key, "no ltla member")
            return {}
        with z.open(member) as fh:
            df = pd.read_csv(_io.TextIOWrapper(fh, "utf-8", errors="ignore"))
    print(f"  [{key}] columns: {[str(c)[:60] for c in df.columns][:8]}")
    code_col = _find_col(df, ["geography code"], contains=True)
    stu_col = next((c for c in df.columns
                    if "l15" in str(c).lower()
                    and "student" in str(c).lower()), None)
    tot_col = next((c for c in df.columns if "total" in str(c).lower()), None)
    if code_col is None or stu_col is None or tot_col is None:
        _warn(key, "couldn't find geography/L15/total columns in TS062")
        _note(key, "unrecognised columns")
        return {}
    for _, r in df.iterrows():
        code = _cell(r, code_col)
        stu, tot = _num(r, stu_col), _num(r, tot_col)
        if code and stu is not None and tot:
            by_code[code.strip()] = {
                "students": int(stu),
                "students_pct": round(100 * stu / tot, 1)}
    gdf = gpd.read_file(bpath)
    gcode = next((c for c in gdf.columns
                  if re.fullmatch(r"lad\d*cd", str(c).lower())), None)
    gname = next((c for c in gdf.columns
                  if re.fullmatch(r"lad\d*nm", str(c).lower())), None)

    def stu_props(f):
        e = by_code.get(_cell(f, gcode)) if gcode else None
        return dict(e) if e else {}

    keep = gdf[gdf.apply(lambda f: bool(stu_props(f)), axis=1)]
    print(f"  [{key}] {len(keep)}/{len(gdf)} authorities matched")
    rows = _emit(keep, key, name_col=gname, id_col=gcode, want="polygon",
                 props_fn=stu_props)
    _note(key, how, len(rows))
    return {key: rows}


def build_student_accom():
    """Existing purpose-built student accommodation stock from OSM
    (building=dormitory / amenity=student_accommodation), reduced to points
    — the PBSA competition map. Extracted by the workflow's osmium step."""
    key = "student_accom"
    path = RAW / "osm_student_accom.geojson"
    if not path.exists():
        _warn(key, "no osm_student_accom.geojson — run via the workflow "
                   "(its osmium step extracts dormitories from the UK PBF)")
        _note(key, "no OSM extract")
        return {}
    print(f"  [{key}] reading {path.name} ...")
    gdf = gpd.read_file(path)
    gdf = gdf[gdf.geometry.notna() & ~gdf.geometry.is_empty].copy()
    # One point per building/site: centroids for polygonal footprints.
    gdf["geometry"] = gdf.geometry.centroid
    name_col = _find_col(gdf, ["name"], contains=False)
    op_col = _find_col(gdf, ["operator"], contains=False)
    rows = _emit(gdf, key, name_col=name_col, want="point", simplify=False,
                 props_fn=lambda f: {k: v for k, v in
                                     (("operator", _cell(f, op_col) if op_col else None),)
                                     if v})
    _note(key, "OSM dormitories", len(rows))
    return {key: rows}


def build_bus_route():
    """OSM route=bus relations -> national bus route lines. The workflow's
    osmium step filters the UK PBF to bus routes and ogr2ogr assembles the
    relation geometry (multilinestrings layer); route number and operator
    are parsed from the OSM other_tags hstore. BODS GTFS adds frequencies
    to the STOPS separately — this layer is the network's shape."""
    key = "bus_route"
    path = RAW / "osm_bus_routes.geojson"
    if not path.exists():
        _warn(key, "no osm_bus_routes.geojson — run via the workflow (its "
                   "osmium step extracts route=bus from the UK PBF)")
        _note(key, "no OSM bus extract")
        return {}
    print(f"  [{key}] reading {path.name} ...")
    gdf = gpd.read_file(path)

    def tagval(s, k):
        if not isinstance(s, str):
            return None
        m = re.search('"' + k + '"=>"([^"]*)"', s)
        return m.group(1) if m else None

    gdf = gdf.copy()
    if "other_tags" in gdf.columns:
        rt = gdf["other_tags"].apply(lambda s: tagval(s, "route"))
        gdf = gdf[(rt == "bus") | rt.isna()]
        gdf["ref"] = gdf["other_tags"].apply(lambda s: tagval(s, "ref"))
        gdf["operator"] = gdf["other_tags"].apply(
            lambda s: tagval(s, "operator"))
    name_col = "name" if "name" in gdf.columns else None
    has_ref = "ref" in gdf.columns
    has_op = "operator" in gdf.columns
    # Routes are enormous multilinestrings (a 40 km route carries thousands
    # of vertices) and the display path simplifies per query — pre-simplify
    # to 25 m here so urban bbox fetches stay ~1 s instead of ~18 s.
    try:
        g27 = gdf.to_crs(27700)
        g27["geometry"] = g27.geometry.simplify(25, preserve_topology=False)
        gdf = g27.to_crs(4326)
        gdf = gdf[gdf.geometry.notna() & ~gdf.geometry.is_empty]
    except Exception as exc:
        print(f"  [{key}] pre-simplify skipped ({exc})")
    rows = _emit(gdf, key, name_col=name_col, want="line", simplify=False,
                 props_fn=lambda f: {k: v for k, v in (
                     ("ref", _cell(f, "ref") if has_ref else None),
                     ("operator", _cell(f, "operator") if has_op else None))
                     if v})
    _note(key, "OSM route=bus relations", len(rows))
    return {key: rows}


PLANIT_BASE = "https://www.planit.org.uk/api/applics/json"


_PLANIT_DIAG = {"n": 0}


def _planit_count(params, sleep_s):
    """One PlanIt count query -> int total, or None on failure. Sleeps
    sleep_s before the request (their fair-use ask), honours 429
    Retry-After (capped), and fails FAST otherwise — a national sweep must
    never turn one bad endpoint into a six-hour hang."""
    url = PLANIT_BASE + "?" + urllib.parse.urlencode(params)
    # The diagnostic run proved PlanIt WORKS from CI but 429s hard (60s
    # Retry-After on most requests) — so patience on 429 is the strategy,
    # while every other failure stays fast-fail.
    for attempt in range(4):
        time.sleep(sleep_s)
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "mastermapper-pipeline/1.0",
                              "Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=20) as r:
                j = json.loads(r.read().decode("utf-8", "replace"))
            for k in ("total", "count", "total_results"):
                v = j.get(k)
                if isinstance(v, int):
                    return v
            print(f"  [planit_rates] unrecognised response keys: "
                  f"{sorted(j.keys())[:8]}")
            return None
        except urllib.error.HTTPError as e:
            if e.code == 429 and attempt < 3:
                wait = min(int(e.headers.get("Retry-After") or "30"), 90)
                time.sleep(wait)
                continue
            if _PLANIT_DIAG["n"] < 5:
                _PLANIT_DIAG["n"] += 1
                print(f"  [planit_rates] HTTP {e.code} from PlanIt "
                      f"({params.get('auth', '?')})")
            return None
        except Exception as exc:
            if _PLANIT_DIAG["n"] < 5:
                _PLANIT_DIAG["n"] += 1
                print(f"  [planit_rates] request failed "
                      f"({params.get('auth', '?')}): {exc}")
            return None
    return None


def build_planit():
    """LPA approval rates from PlanIt (planit.org.uk aggregates every
    council's planning register, OGL-friendly, free with fair-use limits).
    Three COUNT queries per authority (Permitted / Rejected / Withdrawn,
    last 3 years) — no application records are pulled, so a national build
    is ~1,200 tiny requests. Joined onto LAD polygons by name, like HDT.
    PLANIT_SLEEP (seconds between requests, default 1.2) and
    PLANIT_MAX_AUTH (bound for smoke tests) tune the run."""
    key = "planit_rates"
    bpath, _bhow = _resolve_arcgis_source(
        "lad_boundary", ["LAD_BOUNDARY_SRC", "LAD_BOUNDARIES_SRC"],
        LAD_DEFAULT_URL, "lad-boundaries.geojson")
    if bpath is None:
        _warn(key, "no LAD boundaries to join approval rates onto")
        _note(key, "no LAD boundaries")
        return {}
    gdf = gpd.read_file(bpath)
    gcode = next((c for c in gdf.columns
                  if re.fullmatch(r"lad\d*cd", str(c).lower())), None)
    gname = next((c for c in gdf.columns
                  if re.fullmatch(r"lad\d*nm", str(c).lower())), None)
    if gname is None:
        _warn(key, "no LAD name column")
        _note(key, "no LAD name column")
        return {}

    sleep_s = float(os.environ.get("PLANIT_SLEEP") or "1.2")
    max_auth = int(os.environ.get("PLANIT_MAX_AUTH") or "0") or None
    # Hard wall-clock budget: emit whatever is gathered when it runs out,
    # rather than letting a slow/tarpitting endpoint eat the whole job.
    budget_s = float(os.environ.get("PLANIT_TIME_BUDGET_MIN") or "45") * 60
    t0 = time.monotonic()
    # Progress persists across runs (actions/cache in load-planit.yml):
    # already-gathered authorities are skipped, so the daily job walks the
    # country a chunk at a time under PlanIt's rate limit.
    prog_path = Path(os.environ.get("PLANIT_PROGRESS")
                     or (RAW / "planit_progress.json"))
    prior = {}
    if prog_path.exists():
        try:
            prior = json.loads(prog_path.read_text())
            print(f"  [{key}] resuming — {len(prior)} authorities already "
                  "gathered in previous runs")
        except Exception:
            prior = {}
    end = pd.Timestamp.now(tz="UTC").date()
    start = end - pd.Timedelta(days=3 * 365)
    window = {"start_date": start.isoformat(), "end_date": end.isoformat(),
              "pg_sz": 1}

    # England+Wales only (PlanIt coverage): skip Scottish (S...) codes to
    # save ~100 wasted authorities' worth of requests.
    rows_src = gdf[~gdf[gcode].astype(str).str.startswith("S")] \
        if gcode else gdf
    if max_auth:
        rows_src = rows_src.head(max_auth)

    stats, missed = dict(prior), []
    for i, (_, f) in enumerate(rows_src.iterrows()):
        if time.monotonic() - t0 > budget_s:
            print(f"  [{key}] time budget exhausted after {i} authorities — "
                  f"emitting the {len(stats)} gathered so far")
            break
        auth = _cell(f, gname)
        if not auth:
            continue
        if re.sub(r"[^a-z0-9]", "", auth.lower()) in stats:
            continue                    # gathered in a previous run
        got = {}
        for state in ("Permitted", "Rejected"):
            n = _planit_count({**window, "auth": auth, "app_state": state},
                              sleep_s)
            if n is None:
                break
            got[state] = n
        if len(got) < 2 or (got["Permitted"] + got["Rejected"]) == 0:
            missed.append(auth)
            continue
        decided = got["Permitted"] + got["Rejected"]
        stats[re.sub(r"[^a-z0-9]", "", auth.lower())] = {
            "approved_3y": got["Permitted"], "refused_3y": got["Rejected"],
            "approval_pct": round(100 * got["Permitted"] / decided, 1),
            "apps_year": round(decided / 3),
        }
        if (i + 1) % 25 == 0:
            print(f"  [{key}] {i + 1}/{len(rows_src)} authorities queried, "
                  f"{len(stats)} with data, "
                  f"{(time.monotonic() - t0) / 60:.0f} min elapsed")
    if missed:
        print(f"  [{key}] no PlanIt data for {len(missed)} authorities "
              f"(first few: {missed[:6]})")
    try:
        prog_path.parent.mkdir(parents=True, exist_ok=True)
        prog_path.write_text(json.dumps(stats))
        print(f"  [{key}] progress saved: {len(stats)} authorities total")
    except Exception as exc:
        print(f"  [{key}] progress save failed ({exc})")

    def planit_props(f):
        e = stats.get(re.sub(r"[^a-z0-9]", "", _cell(f, gname).lower()))
        return dict(e) if e else {}

    keep = gdf[gdf.apply(lambda f: bool(planit_props(f)), axis=1)]
    print(f"  [{key}] {len(keep)}/{len(gdf)} authorities matched")
    rows = _emit(keep, key, name_col=gname, id_col=gcode, want="polygon",
                 props_fn=planit_props)
    _note(key, "planit.org.uk counts", len(rows))
    return {key: rows}


def build_alc():
    key = "alc"
    path, how = _resolve_arcgis_source(key, ["ALC_SRC"], ALC_DEFAULT_URL,
                                       "alc-england.geojson")
    if path is None:
        _warn(key, f"{how} — set ALC_SRC to the Natural England Provisional"
                   " ALC (England) GeoJSON or FeatureServer query URL")
        _note(key, how)
        return {}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    gdf = gpd.read_file(path)
    grade_col = _find_col(gdf, ["alc_grade", "alcgrade", "alc grade", "grade"],
                          contains=True)
    rows = _emit(gdf, key, name_col=grade_col, want="polygon",
                 props_fn=lambda f: {"alc_grade": _cell(f, grade_col)})
    _note(key, how, len(rows))
    return {key: rows}


def build_water_availability():
    key = "water_availability"
    path, how = _resolve_source(key, ["WATER_AVAILABILITY_SRC", "WATER_SRC"],
                                WATER_DEFAULT_URL, "water-availability.geojson")
    if path is not None:
        # The EA CAMS zip carries six variants; Q95 (low-flow availability) is
        # THE screen for reliable new abstraction, so prefer it.
        path = _maybe_unzip_geo(path, key, prefer=("q95", "cycle_2"))
    if path is None:
        _warn(key, "no source — set WATER_AVAILABILITY_SRC (or WATER_SRC) to"
                   " the EA CAMS 'water resource availability' GeoJSON/SHP"
                   " download (https://environment.data.gov.uk, search: CAMS"
                   " water resource availability).")
        _note(key, "WATER_AVAILABILITY_SRC not set (no default URL)")
        return {}
    print(f"  [{key}] reading {path.name} ({how}) ...")
    gdf = gpd.read_file(path)
    # The last run matched NO status column and every polygon loaded with
    # empty props (flat colour, nothing to say on hover). Log the real column
    # names, try harder for the availability class (Q95 = the low-flow screen
    # that matters), and carry EVERY short scalar attribute so a miss never
    # silently degrades to "no data" again.
    print(f"  [{key}] columns: {list(gdf.columns)[:14]}")
    status_col = _find_col(gdf, ["q95", "colour", "color", "status", "availab",
                                 "class", "restrict"], contains=True)
    name_col = _find_col(gdf, ["cams", "water_body", "waterbody", "wb_name",
                               "name"], contains=True)
    geom_col = gdf.geometry.name

    def _props(feat):
        p = {}
        for c in feat.index:
            if c == geom_col or len(p) >= 10:
                continue
            v = feat[c]
            if v is None:
                continue
            s = str(v).strip()
            if not s or s.lower() == "nan" or len(s) > 90:
                continue
            p[str(c).strip().lower()[:40]] = v if isinstance(v, (int, float)) else s
        if status_col is not None:
            p["status"] = _cell(feat, status_col)
        return p

    rows = _emit(gdf, key, name_col=name_col, want="polygon",
                 assume_epsg=27700,  # EA shapefiles are BNG when CRS missing
                 props_fn=_props)
    _note(key, how, len(rows))
    return {key: rows}


# Build groups: (produces, builder, extra dataset keys that also trigger the
# group even when none of its own outputs were requested — tec_register needs
# the substations built in the same run).
GROUPS = (
    [( [k], (lambda _k=k: build_planning_dataset(_k)), [] )
     for k in PLANNING_DATASETS]
    + [
        (["uni_campus"], build_uni_campus, []),
        (["uni_campus_site", "uni_building"], build_university_group, []),
        (["ptal"], build_ptal, []),
        (["ukpn_sites"], build_ukpn_sites, []),
        (["nged_sites"], build_nged_sites, []),
        (["spen_sites", "npg_sites", "enwl_sites", "ssen_sites"],
         build_dno_group, []),
        (["power_line", "power_substation"], build_power_group,
         ["tec_register"]),
        (["gsp_boundary"], build_gsp_boundary, []),
        (["tec_register"], build_tec_register, []),
        (["lad_boundary", "la_rents"], build_rents_group, []),
        (["alc"], build_alc, []),
        (["water_availability"], build_water_availability, []),
        (["hdt"], build_hdt, []),
        (["planit_rates"], build_planit, []),
        (["bus_route"], build_bus_route, []),
        (["ofcom_fibre"], build_ofcom_fibre, []),
        (["census_students"], build_census_students, []),
        (["student_accom"], build_student_accom, []),
        (["building_height"], build_building_height, []),
        (["build_cost_index"], build_build_cost, []),
        (["lad_income"], build_income, []),
        (["msoa_income"], build_msoa_income, []),
        (["land_value"], build_land_value, []),
        (["cil_rates"], build_cil_rates, []),
        (["lsoa_boundary"], build_lsoa_boundary, []),
        (["local_plan_housing"], build_local_plan_housing, []),
        (["housing_need"], build_housing_need, []),
        (["connectivity_lsoa", "connectivity_lad"], build_connectivity, []),
        (["connectivity_oa"], build_connectivity_oa, []),
    ]
)


def _resolve_datasets(cli_arg):
    raw = cli_arg or os.environ.get("DATASETS", "")
    if not raw.strip():
        return list(ALL_DATASETS)
    wanted = [d.strip() for d in raw.replace(",", " ").split() if d.strip()]
    unknown = [d for d in wanted if d not in ALL_DATASETS]
    if unknown:
        print(f"WARNING: ignoring unknown dataset(s): {', '.join(unknown)}")
    ordered = [d for d in ALL_DATASETS if d in wanted]
    return ordered or list(ALL_DATASETS)


def main() -> int:
    ap = argparse.ArgumentParser(description="Build the map_features import CSV.")
    ap.add_argument("--datasets", help="comma-separated subset of: "
                    + ",".join(ALL_DATASETS) + " (default: all). "
                    "Also settable via the DATASETS env var.")
    ap.add_argument("--out", default=str(OUT), help="output CSV path")
    args = ap.parse_args()

    selected = _resolve_datasets(args.datasets)
    out_path = Path(args.out)
    RAW.mkdir(parents=True, exist_ok=True)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Building datasets: {', '.join(selected)}")
    print(f"  SIMPLIFY_M={SIMPLIFY_M} (polygon simplify tolerance, metres)\n")

    fields = ["dataset", "source_id", "name", "props", "geom_wkt"]
    total = 0
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for produces, builder, also in GROUPS:
            wanted = [d for d in produces if d in selected]
            trigger = wanted or any(d in selected for d in also)
            if not trigger:
                continue
            # Best-effort per group: one failing dataset never aborts the run.
            try:
                built = builder() or {}
            except Exception as exc:
                for d in produces:
                    _warn(d, f"FAILED — skipping: {exc}")
                    _note(d, f"builder failed: {exc}")
                continue
            for d, rows in built.items():
                if d not in selected:
                    # Built only as a dependency (e.g. substations for the TEC
                    # join) — not written to the CSV.
                    print(f"  [{d}] built as a dependency only (not requested)")
                    continue
                for r in rows:
                    w.writerow({
                        "dataset": r["dataset"],
                        "source_id": r["source_id"],
                        "name": r["name"] or "",
                        "props": json.dumps(r["props"], separators=(",", ":"),
                                            ensure_ascii=False),
                        "geom_wkt": r["geom_wkt"],
                    })
                total += len(rows)
                print(f"  [{d}] {len(rows)} features written")

    size_mb = out_path.stat().st_size / 1e6
    print("\nSummary")
    print(f"  output: {out_path}  ({size_mb:.2f} MB, {total} features)")
    for d in selected:
        st = STATUS.get(d)
        if st is None:
            print(f"    {d:20s} (not built)")
        elif st["count"] is None:
            print(f"    {d:20s} skipped — {st['reason']}")
        else:
            print(f"    {d:20s} {st['count']} features ({st['reason']})")
    if total == 0:
        print("\nNo features written. Check the warnings above — most datasets"
              " download by default, but ptal/tec_register/la_rents/"
              "water_availability need operator-supplied *_SRC env vars.")
        return 1
    print("\nNext: python supabase/loaders/load_datasets.py "
          "(needs SUPABASE_URL + SUPABASE_SERVICE_KEY).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
